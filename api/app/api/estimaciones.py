"""Router de estimaciones."""
import base64
import binascii
from datetime import date, datetime, time
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, status
from openpyxl import load_workbook
from pydantic import BaseModel

from app.documents.estimacion import Estimacion, FilaEstimacion
from app.middleware.aplicacion import ContextoAplicacion, contexto_aplicacion, contexto_escritura

router = APIRouter(prefix="/estimaciones", tags=["estimaciones"])


class EstimacionIn(BaseModel):
    requerimiento_id: str | None = None
    titulo: str | None = None
    cliente: str | None = None
    iniciativa: str | None = None
    fecha_estimacion: datetime | None = None
    archivo: str | None = None
    filas: list[FilaEstimacion] = []


class EstimacionUploadIn(BaseModel):
    file_base64: str
    file_name: str = "estimacion.xlsx"


def _valor_fila(fila: FilaEstimacion | dict[str, Any], campo: str, default: Any = 0) -> Any:
    if isinstance(fila, dict):
        return fila.get(campo, default)
    return getattr(fila, campo, default)


def _calcular_totales(filas: list[FilaEstimacion] | list[dict[str, Any]]) -> dict[str, float | int]:
    total_horas = sum(float(_valor_fila(fila, "horas_totales") or 0) for fila in filas)
    return {
        "total_filas": len(filas),
        "total_horas": total_horas,
        "total_horas_estimadas": sum(
            float(_valor_fila(fila, "horas_estimadas") or 0) for fila in filas
        ),
        "total_mejor_caso": sum(float(_valor_fila(fila, "mejor_caso") or 0) for fila in filas),
        "total_peor_caso": sum(float(_valor_fila(fila, "peor_caso") or 0) for fila in filas),
        "total_promedio": sum(float(_valor_fila(fila, "promedio") or 0) for fila in filas),
        "total_horas_finales": total_horas,
    }


def _crear_bucket() -> dict[str, float | int]:
    return {"count": 0, "estimated": 0, "best": 0, "worst": 0, "average": 0, "total": 0}


def _agregar_a_resumen(
    resumen: dict[str, dict[str, float | int]], clave: str, fila: FilaEstimacion
) -> None:
    if clave not in resumen:
        resumen[clave] = _crear_bucket()
    resumen[clave]["count"] += 1
    resumen[clave]["estimated"] += fila.horas_estimadas
    resumen[clave]["best"] += fila.mejor_caso
    resumen[clave]["worst"] += fila.peor_caso
    resumen[clave]["average"] += fila.promedio
    resumen[clave]["total"] += fila.horas_totales


def _construir_summary(estimacion: Estimacion) -> dict[str, dict[str, dict[str, float | int]]]:
    by_type: dict[str, dict[str, float | int]] = {}
    by_sprint: dict[str, dict[str, float | int]] = {}
    by_complexity: dict[str, dict[str, float | int]] = {}
    for fila in estimacion.filas:
        _agregar_a_resumen(by_type, fila.tipo_tarea or "SIN TIPO", fila)
        sprint = str(fila.sprint) if fila.sprint is not None else "SIN SPRINT"
        _agregar_a_resumen(by_sprint, sprint, fila)
        _agregar_a_resumen(by_complexity, fila.complejidad or "SIN COMPLEJIDAD", fila)
    return {"byType": by_type, "bySprint": by_sprint, "byComplexity": by_complexity}


def _celda(row: list[Any], index: int) -> Any:
    return row[index] if len(row) > index else None


def _texto(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _entero(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def _decimal(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def _fecha_excel(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for formato in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, formato)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None
    return None


@router.get("")
async def listar(ctx: ContextoAplicacion = Depends(contexto_aplicacion)):
    return await Estimacion.find(ctx.filtro()).to_list()


@router.get("/por-requerimiento/{requerimiento_id}")
async def obtener_por_requerimiento(
    requerimiento_id: str, ctx: ContextoAplicacion = Depends(contexto_aplicacion)
):
    estimacion = await Estimacion.find_one({**ctx.filtro(), "requerimiento_id": requerimiento_id})
    if estimacion is None:
        return {
            "exists": False,
            "estimacion": None,
            "summary": {"byType": {}, "bySprint": {}, "byComplexity": {}},
        }
    return {"exists": True, "estimacion": estimacion, "summary": _construir_summary(estimacion)}


@router.post("/upload/{requerimiento_id}", status_code=status.HTTP_201_CREATED)
async def upload_estimacion(
    requerimiento_id: str,
    datos: EstimacionUploadIn,
    ctx: ContextoAplicacion = Depends(contexto_escritura),
):
    try:
        buffer = base64.b64decode(datos.file_base64)
    except (ValueError, TypeError, binascii.Error) as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Archivo base64 inválido") from exc

    try:
        wb = load_workbook(BytesIO(buffer), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "No se pudo leer el archivo Excel"
        ) from exc

    try:
        sheet = wb.active
        raw = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        wb.close()

    meta_title = _celda(raw[0], 0) if len(raw) > 0 else None
    meta_client = _celda(raw[1], 1) if len(raw) > 1 else None
    meta_initiative = _celda(raw[2], 1) if len(raw) > 2 else None
    meta_date = _celda(raw[3], 1) if len(raw) > 3 else None

    header_idx = 6
    for index in range(min(len(raw), 15)):
        first_cell = str(_celda(raw[index], 0) or "").lower().strip()
        if first_cell in ("no.", "no"):
            header_idx = index
            break

    filas: list[FilaEstimacion] = []
    for row in raw[header_idx + 1 :]:
        numero_raw = _celda(row, 0)
        if numero_raw in (None, ""):
            continue
        try:
            numero = _entero(numero_raw)
            sprint = _entero(_celda(row, 4))
            fila = FilaEstimacion(
                numero=numero,
                epica_feature=_texto(_celda(row, 1)),
                historia_usuario=_texto(_celda(row, 2)),
                tipo_tarea=_texto(_celda(row, 3)),
                sprint=sprint,
                id_epm=_texto(_celda(row, 5)),
                id_hitss=_texto(_celda(row, 6)),
                actividad=_texto(_celda(row, 7)),
                complejidad=_texto(_celda(row, 8)),
                horas_estimadas=_decimal(_celda(row, 9)),
                mejor_caso=_decimal(_celda(row, 10)),
                peor_caso=_decimal(_celda(row, 11)),
                promedio=_decimal(_celda(row, 12)),
                metodologia_10=_decimal(_celda(row, 13)),
                horas_totales=_decimal(_celda(row, 14)),
            )
        except (TypeError, ValueError):
            continue
        if fila.numero is None:
            continue
        filas.append(fila)

    if not filas:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "No se encontraron filas válidas en el Excel"
        )

    existing = await Estimacion.find_one(
        {"requerimiento_id": requerimiento_id, "aplicacion_id": ctx.codigo}
    )
    if existing is not None:
        await existing.delete()

    now = datetime.utcnow()
    estimacion = Estimacion(
        aplicacion_id=ctx.codigo,
        requerimiento_id=requerimiento_id,
        titulo=_texto(meta_title),
        cliente=_texto(meta_client),
        iniciativa=_texto(meta_initiative),
        fecha_estimacion=_fecha_excel(meta_date),
        archivo=datos.file_name,
        subido_en=now,
        filas=filas,
        **_calcular_totales(filas),
    )
    await estimacion.insert()
    return estimacion


@router.get("/{estimacion_id}")
async def obtener(
    estimacion_id: str, ctx: ContextoAplicacion = Depends(contexto_aplicacion)
):
    estimacion = await Estimacion.get(estimacion_id)
    if estimacion is None or estimacion.aplicacion_id not in ctx.codigos:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estimación no encontrada")
    return estimacion


@router.post("", status_code=status.HTTP_201_CREATED)
async def crear(datos: EstimacionIn, ctx: ContextoAplicacion = Depends(contexto_escritura)):
    estimacion = Estimacion(
        aplicacion_id=ctx.codigo,
        **datos.model_dump(),
        **_calcular_totales(datos.filas),
    )
    await estimacion.insert()
    return estimacion


@router.put("/{estimacion_id}")
async def actualizar(
    estimacion_id: str,
    datos: EstimacionIn,
    ctx: ContextoAplicacion = Depends(contexto_escritura),
):
    estimacion = await Estimacion.get(estimacion_id)
    if estimacion is None or estimacion.aplicacion_id != ctx.codigo:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estimación no encontrada")
    patch = datos.model_dump(exclude_unset=True)
    if "filas" in patch:
        patch.update(_calcular_totales(patch["filas"]))
    await estimacion.set(patch)
    return estimacion


@router.delete("/{estimacion_id}", status_code=status.HTTP_204_NO_CONTENT)
async def eliminar(
    estimacion_id: str, ctx: ContextoAplicacion = Depends(contexto_escritura)
) -> None:
    estimacion = await Estimacion.get(estimacion_id)
    if estimacion is None or estimacion.aplicacion_id != ctx.codigo:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estimación no encontrada")
    await estimacion.delete()


# ── Creación de tareas en Azure DevOps (HITSS / EPM) ──

def _entero_id(valor: str | None) -> int | None:
    if valor in (None, ""):
        return None
    try:
        return int(str(valor).strip())
    except (TypeError, ValueError):
        return None


def _campos_tarea(fila: FilaEstimacion, parent_wi: dict | None = None) -> dict:
    horas = float(fila.horas_totales or fila.metodologia_10 or fila.horas_estimadas or 0)
    today = datetime.utcnow().isoformat() + "Z"
    titulo = fila.actividad or f"Tarea #{fila.numero or ''}".strip()
    campos: dict[str, Any] = {
        "System.Title": titulo,
        "System.Description": titulo,
        "Microsoft.VSTS.Common.Activity": _mapear_actividad(fila.tipo_tarea),
        "Microsoft.VSTS.Scheduling.OriginalEstimate": horas if horas > 0 else 0,
        "Microsoft.VSTS.Scheduling.CompletedWork": 0,
        "Microsoft.VSTS.Scheduling.RemainingWork": horas if horas > 0 else 0,
        "Microsoft.VSTS.Scheduling.StartDate": today,
        "Microsoft.VSTS.Scheduling.FinishDate": today,
    }
    # Heredar AreaPath e IterationPath del padre
    if parent_wi and parent_wi.get("fields"):
        pf = parent_wi["fields"]
        if pf.get("System.AreaPath"):
            campos["System.AreaPath"] = pf["System.AreaPath"]
        if pf.get("System.IterationPath"):
            campos["System.IterationPath"] = pf["System.IterationPath"]
    return campos


_ACTIVITY_MAP = {
    "DESARROLLO": "Development",
    "PRUEBAS": "Testing",
    "QA": "Testing",
    "DESPLIEGUE": "Deployment",
    "INVESTIGACION": "Design",
    "INVESTIGACIÓN": "Design",
    "DOCUMENTACION": "Documentation",
    "DOCUMENTACIÓN": "Documentation",
    "DISEÑO": "Design",
    "REQUISITOS": "Requirements",
}


def _mapear_actividad(tipo_tarea: str | None) -> str:
    if not tipo_tarea:
        return "Development"
    return _ACTIVITY_MAP.get(tipo_tarea.upper(), "Development")


async def _cargar_estimacion(estimacion_id: str, ctx: ContextoAplicacion) -> Estimacion:
    estimacion = await Estimacion.get(estimacion_id)
    if estimacion is None or estimacion.aplicacion_id != ctx.codigo:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estimación no encontrada")
    return estimacion


@router.post("/{estimacion_id}/crear-tareas-hitss")
async def crear_tareas_hitss(
    estimacion_id: str, ctx: ContextoAplicacion = Depends(contexto_escritura)
):
    """Crea la jerarquía Feature → User Story → Task en la org HITSS.

    Replica la lógica de workload-manager:
    1. Agrupa filas por ``historia_usuario``.
    2. Si la fila tiene ``id_hitss``, intenta usarlo como padre directo.
    3. Si no hay padre, busca el Feature por título (``estimacion.iniciativa``)
       y crea/busca la User Story bajo él.
    4. Crea las Tasks heredando AreaPath e IterationPath del padre.
    5. Reintentos automáticos para campos requeridos/inválidos.
    """
    from app.services.azdo_sync import crear_servicio_azdo, leer_config_azdo

    estimacion = await _cargar_estimacion(estimacion_id, ctx)
    try:
        svc = await crear_servicio_azdo(ctx.codigo, "azdo_")
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc)) from exc
    proyecto = await leer_config_azdo(ctx.codigo, "azdo_default_project")
    if not proyecto:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Falta configurar 'azdo_default_project' (proyecto HITSS por defecto)",
        )

    wit_types = await svc.obtener_tipos_work_item(proyecto)
    today = datetime.utcnow().isoformat() + "Z"
    feature_name = (estimacion.iniciativa or "").strip()

    creadas = 0
    errores: list[str] = []

    # Agrupar filas por historia_usuario (como workload-manager)
    groups: dict[str, dict] = {}
    for fila in estimacion.filas:
        key = (fila.historia_usuario or "").strip() or f"__row_{fila.numero}"
        if key not in groups:
            groups[key] = {
                "id_hitss": fila.id_hitss,
                "historia_usuario": fila.historia_usuario,
                "filas": [],
            }
        groups[key]["filas"].append(fila)
        # Usar primer id_hitss no-nulo del grupo
        if fila.id_hitss and not groups[key]["id_hitss"]:
            groups[key]["id_hitss"] = fila.id_hitss

    for hu_name, group in groups.items():
        parent_id: int | None = None
        parent_wi: dict | None = None

        # 1) Intentar usar id_hitss como padre directo
        id_hitss = _entero_id(group["id_hitss"])
        if id_hitss:
            wi = await svc.obtener_work_item(id_hitss)
            if wi:
                parent_id = wi.get("id")
                parent_wi = wi

        # 2) Si no hay padre, buscar/crear HU bajo el Feature
        if not parent_id:
            if not feature_name:
                errores.append(
                    f"HU '{hu_name}': Feature no encontrada "
                    f"(sin nombre de iniciativa en la estimación)"
                )
                continue

            # Buscar Feature existente por título
            feature = await svc.buscar_work_item_por_titulo(
                proyecto, feature_name, wit_types["feature"]
            )
            if not feature:
                # Crear el Feature si no existe
                try:
                    feature = await svc.crear_work_item(
                        proyecto,
                        wit_types["feature"],
                        {"System.Title": feature_name},
                    )
                except RuntimeError as exc:
                    errores.append(f"Feature '{feature_name}': {exc}")
                    continue

            feature_id = feature.get("id")
            if feature_id and not estimacion.created_feature_hitss:
                estimacion.created_feature_hitss = int(feature_id)

            # Buscar HU existente, si no existe crearla
            hu_title = (group["historia_usuario"] or hu_name).strip()
            hu = await svc.buscar_work_item_por_titulo(
                proyecto, hu_title, wit_types["userStory"]
            )
            if not hu:
                hu_fields: dict[str, Any] = {
                    "System.Title": hu_title,
                    "System.Description": f"<div>{hu_title}</div>",
                    "Microsoft.VSTS.Common.AcceptanceCriteria": f"<div>{hu_title}</div>",
                    "Microsoft.VSTS.Scheduling.StartDate": today,
                    "Microsoft.VSTS.Scheduling.FinishDate": today,
                }
                # Heredar AreaPath/IterationPath del Feature
                if feature.get("fields"):
                    ff = feature["fields"]
                    if ff.get("System.AreaPath"):
                        hu_fields["System.AreaPath"] = ff["System.AreaPath"]
                    if ff.get("System.IterationPath"):
                        hu_fields["System.IterationPath"] = ff["System.IterationPath"]
                try:
                    hu = await svc.crear_work_item(
                        proyecto, wit_types["userStory"], hu_fields, feature_id
                    )
                except RuntimeError as exc:
                    errores.append(f"HU '{hu_title}': {exc}")
                    continue

            parent_id = hu.get("id")
            parent_wi = hu

            # Guardar ID de HU en las filas del grupo
            if parent_id:
                for fila in group["filas"]:
                    fila.created_hu_hitss = int(parent_id)

        # 3) Crear Tasks bajo el padre
        for fila in group["filas"]:
            if fila.created_task_hitss:
                continue
            try:
                wi = await svc.crear_work_item(
                    proyecto,
                    wit_types["task"],
                    _campos_tarea(fila, parent_wi),
                    parent_id,
                )
                fila.created_task_hitss = int(wi.get("id") or 0) or None
                if fila.created_task_hitss:
                    creadas += 1
            except RuntimeError as exc:
                errores.append(f"Fila {fila.numero}: {exc}")

    estimacion.marcar_actualizado()
    await estimacion.save()
    return {
        "creadas": creadas,
        "errores": errores,
        "estimacion": estimacion,
        "summary": _construir_summary(estimacion),
    }


@router.post("/{estimacion_id}/crear-tareas-epm")
async def crear_tareas_epm(
    estimacion_id: str, ctx: ContextoAplicacion = Depends(contexto_escritura)
):
    """Crea las tareas en la org EPM. Cada fila debe traer ``id_epm`` (HU padre)."""
    from app.services.azdo_sync import crear_servicio_azdo

    estimacion = await _cargar_estimacion(estimacion_id, ctx)
    try:
        svc = await crear_servicio_azdo(ctx.codigo, "azdo2_")
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc)) from exc

    # Cache de proyecto y WI padre por id_epm.
    proyecto_por_epm: dict[int, str] = {}
    padre_wi_cache: dict[int, dict] = {}
    creadas = 0
    errores: list[str] = []

    for fila in estimacion.filas:
        if fila.created_task_epm:
            continue
        parent = _entero_id(fila.id_epm)
        if parent is None:
            errores.append(f"Fila {fila.numero}: sin ID HU EPM")
            continue
        try:
            if parent not in proyecto_por_epm:
                padre = await svc.obtener_work_item(parent)
                if padre is None:
                    errores.append(f"Fila {fila.numero}: HU EPM {parent} no encontrada")
                    continue
                proyecto = (padre.get("fields") or {}).get("System.TeamProject")
                if not proyecto:
                    errores.append(
                        f"Fila {fila.numero}: no se pudo determinar el proyecto de la HU {parent}"
                    )
                    continue
                proyecto_por_epm[parent] = proyecto
                padre_wi_cache[parent] = padre
            wi = await svc.crear_work_item(
                proyecto_por_epm[parent], "Task",
                _campos_tarea(fila, padre_wi_cache.get(parent)), parent
            )
            fila.created_task_epm = int(wi.get("id") or 0) or None
            if fila.created_task_epm:
                creadas += 1
        except RuntimeError as exc:
            errores.append(f"Fila {fila.numero}: {exc}")

    estimacion.marcar_actualizado()
    await estimacion.save()
    return {
        "creadas": creadas,
        "errores": errores,
        "estimacion": estimacion,
        "summary": _construir_summary(estimacion),
    }

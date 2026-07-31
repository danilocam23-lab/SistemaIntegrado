"""Router de importación del Excel 'BITÁCORA GENERAL'."""
from datetime import datetime
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from app.documents.aplicacion import Aplicacion
from app.documents.enums import EstadoEntrega, EstadoRequerimiento
from app.documents.persona import Persona
from app.documents.requerimiento import Requerimiento
from app.documents.squad import Squad
from app.documents.tarifa import Tarifa
from app.importer.excel_importer import ImportadorExcel
from app.middleware.aplicacion import (
    ContextoAplicacion,
    contexto_aplicacion,
)

router = APIRouter(prefix="/importacion", tags=["importacion"])


@router.post("/excel")
async def importar_excel(
    archivo: UploadFile = File(...),
    hoja: str | None = None,
    ctx: ContextoAplicacion = Depends(contexto_aplicacion),
) -> dict:
    """Importa el Excel 'BITÁCORA GENERAL' a la aplicación activa.

    Cada fila es un requerimiento + una entrega; los catálogos (personas,
    squads, tarifas) se crean si no existen.
    """
    nombre = (archivo.filename or "").lower()
    if not nombre.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "El archivo debe ser .xlsx")

    contenido = await archivo.read()
    try:
        if ctx.modo_consolidado:
            resultado = await _importar_consolidado(ctx, contenido, hoja)
        else:
            resultado = await ImportadorExcel(ctx.codigo, contenido, hoja).ejecutar()
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, f"No se pudo importar el archivo: {exc}"
        ) from exc

    return {
        "filas_procesadas": resultado.filas_procesadas,
        "requerimientos_creados": resultado.requerimientos_creados,
        "requerimientos_actualizados": resultado.requerimientos_actualizados,
        "entregas_creadas": resultado.entregas_creadas,
        "entregas_actualizadas": resultado.entregas_actualizadas,
        "festivos_cargados": resultado.festivos_cargados,
        "errores": resultado.errores,
    }


@router.get("/excel/plantilla")
async def exportar_plantilla(
    ctx: ContextoAplicacion = Depends(contexto_aplicacion),
):
    """Exporta plantilla con hojas separadas: REQUERIMIENTOS y ENTREGAS."""
    reqs = await Requerimiento.find(ctx.filtro()).sort("codigo_req").to_list()
    squads = {
        str(doc.id): doc for doc in await Squad.find(ctx.filtro()).to_list()
    }
    personas = {
        str(doc.id): doc for doc in await Persona.find(ctx.filtro()).to_list()
    }
    tarifas_q = {
        "$or": [
            {"aplicacion_id": {"$in": ctx.codigos}},
            {"aplicacion_id": "global"},
        ]
    }
    tarifas_lista = await Tarifa.find(tarifas_q).to_list()
    tarifas = {str(doc.id): doc for doc in tarifas_lista}
    apps = await Aplicacion.find({"codigo": {"$in": ctx.codigos}}).to_list()
    squads_por_nombre = {s.nombre.strip().lower(): s for s in squads.values() if s.nombre}
    personas_por_nombre = {p.nombre.strip().lower(): p for p in personas.values() if p.nombre}
    tarifas_por_anio_tecnologia = {
        (t.anio, (t.ramificacion or "").strip().lower()): t
        for t in tarifas_lista
        if t.ramificacion
    }

    libro = openpyxl.Workbook()
    hoja_req = libro.active
    hoja_req.title = "REQUERIMIENTOS"
    headers_req = [
        "CODIGO DE SOLICITUD DE COTIZACION",
        "COD. DEL REQ",
        "NOMBRE DEL REQUERIMIENTO",
        "ESTADO DE REQUERIMIENTOS",
        "FECHA Y HORA DE SOLICITUD",
        "Squad",
        "LT HITSS",
        "LT EPM",
        "SCRUM",
        "TIPO DE COSTO",
        "TOTAL HORAS ESTIMADAS",
        "FECHA REAL ENTREGA DE ESTIMACIONES",
        "ACTA DE TRABAJO",
        "SEGUIMIENTO",
    ]
    hoja_req.append(headers_req)
    hoja_req.freeze_panes = "A2"
    hoja_req.auto_filter.ref = f"A1:{get_column_letter(len(headers_req))}1"
    _marcar_requeridas(
        hoja_req,
        headers_req,
        {"COD. DEL REQ", "NOMBRE DEL REQUERIMIENTO", "ESTADO DE REQUERIMIENTOS", "CODIGO DE SOLICITUD DE COTIZACION"},
    )

    hoja_ent = libro.create_sheet("ENTREGAS")
    headers_ent = [
        "COD. DEL REQ",
        "CODIGO DE SOLICITUD DE COTIZACION",
        "CANT. DE ENTREGAS",
        "HORAS A ENTREGAR",
        "FECHA COMPROMETIDA DE ENTREGAS DLLO",
        "F REAL",
        "OBSERVACIONES",
        "MES APROBACIÓN",
        "GARANTIA",
        "ESTADO ENTREGAS",
    ]
    hoja_ent.append(headers_ent)
    hoja_ent.freeze_panes = "A2"
    hoja_ent.auto_filter.ref = f"A1:{get_column_letter(len(headers_ent))}1"
    _marcar_requeridas(
        hoja_ent,
        headers_ent,
        {
            "COD. DEL REQ",
            "CODIGO DE SOLICITUD DE COTIZACION",
            "CANT. DE ENTREGAS",
            "HORAS A ENTREGAR",
            "FECHA COMPROMETIDA DE ENTREGAS DLLO",
            "ESTADO ENTREGAS",
        },
    )

    for req in reqs:
        sol = req.solicitud
        squad = _resolver_catalogo(sol.squad_id, squads, squads_por_nombre)
        lt_hitss = _resolver_catalogo(sol.lt_hitss_id, personas, personas_por_nombre)
        lt_epm = _resolver_catalogo(sol.lt_epm_id, personas, personas_por_nombre)
        scrum = _resolver_catalogo(sol.scrum_id, personas, personas_por_nombre)
        tarifa = tarifas.get(sol.tarifa_id or "")
        if tarifa is None:
            tarifa = tarifas_por_anio_tecnologia.get(
                (sol.anio_tarifa or 0, (sol.tecnologia or "").strip().lower())
            )
        hoja_req.append(
            [
                sol.codigo_sc,
                req.codigo_req,
                req.nombre,
                req.estado,
                req.fecha_solicitud_acta,
                squad.nombre if squad else (sol.squad_id or None),
                " ".join(lt_hitss.nombre.split()) if lt_hitss else (sol.lt_hitss_id or None),
                " ".join(lt_epm.nombre.split()) if lt_epm else (sol.lt_epm_id or None),
                " ".join(scrum.nombre.split()) if scrum else (sol.scrum_id or None),
                sol.tipo_costo,
                req.total_horas_estimadas,
                req.fecha_real_entrega_estimacion,
                req.acta_trabajo,
                req.seguimiento,
            ]
        )

        filas_entrega = req.entregas if req.entregas else []
        for entrega in filas_entrega:
            hoja_ent.append(
                [
                    req.codigo_req,
                    sol.codigo_sc,
                    entrega.numero,
                    entrega.horas,
                    entrega.fecha_comprometida,
                    entrega.fecha_recepcion,
                    entrega.observaciones,
                    entrega.mes_aprobacion,
                    "SI" if entrega.garantia else None,
                    entrega.estado,
                ]
            )

    _autoajustar_columnas(hoja_req)
    _autoajustar_columnas(hoja_ent)
    _crear_hoja_instrucciones(libro)
    _crear_hoja_catalogos(
        libro=libro,
        hoja_req=hoja_req,
        headers_req=headers_req,
        hoja_ent=hoja_ent,
        headers_ent=headers_ent,
        apps=apps,
        squads=squads,
        personas=personas,
    )

    contenido = BytesIO()
    libro.save(contenido)
    contenido.seek(0)
    codigo_archivo = "consolidado" if ctx.modo_consolidado else ctx.codigo
    nombre = f"plantilla_requerimientos_entregas_{codigo_archivo}.xlsx"
    return StreamingResponse(
        contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@router.post("/excel/previsualizar")
async def previsualizar_importacion(
    archivo: UploadFile = File(...),
    hoja: str | None = None,
    ctx: ContextoAplicacion = Depends(contexto_aplicacion),
) -> dict:
    """Valida el archivo y devuelve resumen de registros nuevos/actualizados."""
    nombre = (archivo.filename or "").lower()
    if not nombre.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "El archivo debe ser .xlsx")
    contenido = await archivo.read()
    try:
        if ctx.modo_consolidado:
            return await _previsualizar_consolidado(ctx, contenido, hoja)
        return await ImportadorExcel(ctx.codigo, contenido, hoja).previsualizar()
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"No se pudo previsualizar el archivo: {exc}",
        ) from exc


def _mes_a_texto(fecha: datetime | None) -> str | None:
    if fecha is None:
        return None
    meses = (
        "enero",
        "febrero",
        "marzo",
        "abril",
        "mayo",
        "junio",
        "julio",
        "agosto",
        "septiembre",
        "octubre",
        "noviembre",
        "diciembre",
    )
    return meses[fecha.month - 1]


def _resolver_catalogo(
    valor: str | None,
    por_id: dict,
    por_texto: dict,
):
    if not valor:
        return None
    doc = por_id.get(valor)
    if doc is not None:
        return doc
    return por_texto.get(valor.strip().lower())


def _marcar_requeridas(hoja, headers: list[str], requeridas: set[str]) -> None:
    color_requerida = PatternFill("solid", fgColor="FEE2E2")
    color_opcional = PatternFill("solid", fgColor="F1F5F9")
    for i, header in enumerate(headers, start=1):
        celda = hoja.cell(row=1, column=i)
        celda.font = Font(bold=True)
        celda.fill = color_requerida if header in requeridas else color_opcional


def _autoajustar_columnas(hoja) -> None:
    max_ancho = 45
    for col in range(1, hoja.max_column + 1):
        letra = get_column_letter(col)
        ancho = 0
        for row in range(1, min(hoja.max_row, 150) + 1):
            valor = hoja.cell(row=row, column=col).value
            if valor is None:
                continue
            ancho = max(ancho, len(str(valor)))
        hoja.column_dimensions[letra].width = min(max(ancho + 2, 14), max_ancho)


def _crear_hoja_instrucciones(libro) -> None:
    ws = libro.create_sheet("INSTRUCCIONES")
    ws.append(["Plantilla de importación masiva"])
    ws.append(
        [
            "Use las hojas REQUERIMIENTOS y ENTREGAS. Las columnas en rojo son requeridas. "
            "Las grises son opcionales."
        ]
    )
    ws.append(["No agregue columnas nuevas ni cambie los encabezados."])
    ws.append(
        [
            "Las columnas calculadas automáticamente (ANS, % de entrega) se excluyen "
            "de esta plantilla para evitar inconsistencias."
        ]
    )
    ws.append(
        [
            "Para campos tipo lista (estados, garantía, mes), "
            "use el desplegable."
        ]
    )
    ws.append(
        [
            "En la hoja ENTREGAS la clave única es: COD. DEL REQ + CODIGO DE SOLICITUD DE COTIZACION."
        ]
    )
    ws["A1"].font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 120


def _crear_hoja_catalogos(
    *,
    libro,
    hoja_req,
    headers_req: list[str],
    hoja_ent,
    headers_ent: list[str],
    apps: list,
    squads: dict,
    personas: dict,
) -> None:
    ws = libro.create_sheet("CATALOGOS")
    columnas = [
        ("ESTADOS_REQ", sorted([x.value for x in EstadoRequerimiento])),
        ("ESTADOS_ENT", sorted([x.value for x in EstadoEntrega])),
        ("TIPO_COSTO", ["FIJO", "TIEMPO Y MATERIALES"]),
        ("GARANTIA", ["SI", "NO"]),
        (
            "MESES",
            [
                "enero",
                "febrero",
                "marzo",
                "abril",
                "mayo",
                "junio",
                "julio",
                "agosto",
                "septiembre",
                "octubre",
                "noviembre",
                "diciembre",
            ],
        ),
        ("LT_HITSS", sorted({p.nombre for p in personas.values() if getattr(p, "rol_operativo", "") == "LT_HITSS"})),
        ("LT_EPM", sorted({p.nombre for p in personas.values() if getattr(p, "rol_operativo", "") == "LT_EPM"})),
        ("SCRUMS", sorted({p.nombre for p in personas.values() if getattr(p, "rol_operativo", "") == "SCRUM"})),
        ("SQUADS", sorted({s.nombre for s in squads.values()} | {a.nombre for a in apps if a.nombre})),
    ]

    for col_idx, (titulo, valores) in enumerate(columnas, start=1):
        ws.cell(row=1, column=col_idx, value=titulo).font = Font(bold=True)
        fila = 2
        for v in valores:
            if not v:
                continue
            ws.cell(row=fila, column=col_idx, value=v)
            fila += 1
        ws.column_dimensions[get_column_letter(col_idx)].width = 28

    rangos = {}
    for idx, (titulo, _) in enumerate(columnas, start=1):
        letra = get_column_letter(idx)
        ultima = max(2, ws.max_row)
        rangos[titulo] = f"CATALOGOS!${letra}$2:${letra}${ultima}"
    nombres = {
        "ESTADOS_REQ": "LISTA_ESTADOS_REQ",
        "ESTADOS_ENT": "LISTA_ESTADOS_ENT",
        "TIPO_COSTO": "LISTA_TIPO_COSTO",
        "GARANTIA": "LISTA_GARANTIA",
        "MESES": "LISTA_MESES",
        "LT_HITSS": "LISTA_LT_HITSS",
        "LT_EPM": "LISTA_LT_EPM",
        "SCRUMS": "LISTA_SCRUMS",
        "SQUADS": "LISTA_SQUADS",
    }
    for clave, nombre_rango in nombres.items():
        _definir_nombre(libro, nombre_rango, rangos[clave])

    por_columna_req = {
        "ESTADO DE REQUERIMIENTOS": "LISTA_ESTADOS_REQ",
        "TIPO DE COSTO": "LISTA_TIPO_COSTO",
        "LT HITSS": "LISTA_LT_HITSS",
        "LT EPM": "LISTA_LT_EPM",
        "SCRUM": "LISTA_SCRUMS",
        "Squad": "LISTA_SQUADS",
    }
    por_columna_ent = {
        "ESTADO ENTREGAS": "LISTA_ESTADOS_ENT",
        "GARANTIA": "LISTA_GARANTIA",
        "MES APROBACIÓN": "LISTA_MESES",
    }
    _aplicar_validaciones(hoja_req, headers_req, por_columna_req)
    _aplicar_validaciones(hoja_ent, headers_ent, por_columna_ent)
    ws.sheet_state = "hidden"


def _aplicar_validaciones(hoja, headers: list[str], por_columna: dict[str, str]) -> None:
    columnas = {h: i + 1 for i, h in enumerate(headers)}
    fila_inicio, fila_fin = 2, 5000
    for campo, nombre_rango in por_columna.items():
        col = columnas.get(campo)
        if not col:
            continue
        dv = DataValidation(type="list", formula1=f"={nombre_rango}", allow_blank=True)
        hoja.add_data_validation(dv)
        letra = get_column_letter(col)
        dv.add(f"{letra}{fila_inicio}:{letra}{fila_fin}")


def _definir_nombre(libro, nombre: str, referencia: str) -> None:
    nombres = libro.defined_names
    definido = DefinedName(name=nombre, attr_text=referencia)
    if hasattr(nombres, "add"):
        nombres.add(definido)
        return
    if hasattr(nombres, "append"):
        nombres.append(definido)
        return
    if hasattr(nombres, "definedName"):
        nombres.definedName.append(definido)
        return
    raise ValueError("No se pudo registrar los rangos del catálogo en el archivo Excel.")


async def _importar_consolidado(ctx: ContextoAplicacion, contenido: bytes, hoja: str | None) -> object:
    apps = await Aplicacion.find({"codigo": {"$in": ctx.codigos}}).to_list()
    archivos_por_app = _separar_archivo_por_aplicacion(contenido, apps, hoja)
    resultado_total = {
        "filas_procesadas": 0,
        "requerimientos_creados": 0,
        "requerimientos_actualizados": 0,
        "entregas_creadas": 0,
        "entregas_actualizadas": 0,
        "festivos_cargados": 0,
        "errores": [],
    }
    for codigo, archivo_app in archivos_por_app.items():
        res = await ImportadorExcel(codigo, archivo_app, hoja).ejecutar()
        resultado_total["filas_procesadas"] += res.filas_procesadas
        resultado_total["requerimientos_creados"] += res.requerimientos_creados
        resultado_total["requerimientos_actualizados"] += res.requerimientos_actualizados
        resultado_total["entregas_creadas"] += res.entregas_creadas
        resultado_total["entregas_actualizadas"] += res.entregas_actualizadas
        resultado_total["festivos_cargados"] += res.festivos_cargados
        resultado_total["errores"].extend([f"[{codigo}] {e}" for e in res.errores])

    class _Resultado:
        def __init__(self, data: dict):
            self.filas_procesadas = data["filas_procesadas"]
            self.requerimientos_creados = data["requerimientos_creados"]
            self.requerimientos_actualizados = data["requerimientos_actualizados"]
            self.entregas_creadas = data["entregas_creadas"]
            self.entregas_actualizadas = data["entregas_actualizadas"]
            self.festivos_cargados = data["festivos_cargados"]
            self.errores = data["errores"]

    return _Resultado(resultado_total)


async def _previsualizar_consolidado(ctx: ContextoAplicacion, contenido: bytes, hoja: str | None) -> dict:
    apps = await Aplicacion.find({"codigo": {"$in": ctx.codigos}}).to_list()
    archivos_por_app = _separar_archivo_por_aplicacion(contenido, apps, hoja)
    total = {
        "filas_requerimientos": 0,
        "filas_entregas": 0,
        "requerimientos_nuevos": 0,
        "requerimientos_actualizados": 0,
        "entregas_nuevas": 0,
        "entregas_actualizadas": 0,
        "detalle_requerimientos_nuevos": [],
        "detalle_requerimientos_actualizados": [],
        "detalle_entregas_nuevas": [],
        "detalle_entregas_actualizadas": [],
    }
    for codigo, archivo_app in archivos_por_app.items():
        res = await ImportadorExcel(codigo, archivo_app, hoja).previsualizar()
        total["filas_requerimientos"] += int(res.get("filas_requerimientos", 0))
        total["filas_entregas"] += int(res.get("filas_entregas", 0))
        total["requerimientos_nuevos"] += int(res.get("requerimientos_nuevos", 0))
        total["requerimientos_actualizados"] += int(res.get("requerimientos_actualizados", 0))
        total["entregas_nuevas"] += int(res.get("entregas_nuevas", 0))
        total["entregas_actualizadas"] += int(res.get("entregas_actualizadas", 0))
        total["detalle_requerimientos_nuevos"].extend(
            [f"[{codigo}] {x}" for x in res.get("detalle_requerimientos_nuevos", [])]
        )
        total["detalle_requerimientos_actualizados"].extend(
            [f"[{codigo}] {x}" for x in res.get("detalle_requerimientos_actualizados", [])]
        )
        total["detalle_entregas_nuevas"].extend(
            [f"[{codigo}] {x}" for x in res.get("detalle_entregas_nuevas", [])]
        )
        total["detalle_entregas_actualizadas"].extend(
            [f"[{codigo}] {x}" for x in res.get("detalle_entregas_actualizadas", [])]
        )
    total["detalle_requerimientos_nuevos"] = sorted(set(total["detalle_requerimientos_nuevos"]))[:200]
    total["detalle_requerimientos_actualizados"] = sorted(set(total["detalle_requerimientos_actualizados"]))[:200]
    total["detalle_entregas_nuevas"] = sorted(set(total["detalle_entregas_nuevas"]))[:400]
    total["detalle_entregas_actualizadas"] = sorted(set(total["detalle_entregas_actualizadas"]))[:400]
    return total


def _separar_archivo_por_aplicacion(contenido: bytes, apps: list[Aplicacion], hoja: str | None) -> dict[str, bytes]:
    wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
    app_por_clave: dict[str, str] = {}
    for app in apps:
        app_por_clave[app.codigo.strip().lower()] = app.codigo
        app_por_clave[app.nombre.strip().lower()] = app.codigo

    if "REQUERIMIENTOS" in wb.sheetnames and "ENTREGAS" in wb.sheetnames:
        return _separar_formato_dos_hojas(wb, app_por_clave)
    return _separar_formato_legacy(wb, app_por_clave, hoja)


def _separar_formato_dos_hojas(wb, app_por_clave: dict[str, str]) -> dict[str, bytes]:
    ws_req = wb["REQUERIMIENTOS"]
    ws_ent = wb["ENTREGAS"]
    headers_req = _headers_hoja(ws_req)
    headers_ent = _headers_hoja(ws_ent)
    idx_req = {h: i for i, h in enumerate(headers_req)}
    idx_ent = {h: i for i, h in enumerate(headers_ent)}
    if "Squad" not in idx_req:
        raise ValueError("La hoja REQUERIMIENTOS debe incluir la columna 'Squad' para importar en modo consolidado.")

    req_por_app: dict[str, list[tuple]] = {}
    ent_por_app: dict[str, list[tuple]] = {}
    app_por_req_sc: dict[tuple[str, str], str] = {}

    for row in ws_req.iter_rows(min_row=2, values_only=True):
        codigo_req = _txt(row[idx_req.get("COD. DEL REQ")]) if idx_req.get("COD. DEL REQ") is not None else None
        if not codigo_req:
            continue
        codigo_sc = _txt(row[idx_req.get("CODIGO DE SOLICITUD DE COTIZACION")]) if idx_req.get("CODIGO DE SOLICITUD DE COTIZACION") is not None else None
        codigo_sc = codigo_sc or codigo_req
        squad = _txt(row[idx_req["Squad"]])
        app_codigo = _resolver_app_por_squad(squad, app_por_clave)
        if app_codigo is None:
            raise ValueError(f"No se pudo resolver la aplicación del requerimiento '{codigo_req}' con Squad='{squad or ''}'.")
        req_por_app.setdefault(app_codigo, []).append(row)
        app_por_req_sc[(codigo_req, codigo_sc)] = app_codigo

    for row in ws_ent.iter_rows(min_row=2, values_only=True):
        codigo_req = _txt(row[idx_ent.get("COD. DEL REQ")]) if idx_ent.get("COD. DEL REQ") is not None else None
        if not codigo_req:
            continue
        codigo_sc = _txt(row[idx_ent.get("CODIGO DE SOLICITUD DE COTIZACION")]) if idx_ent.get("CODIGO DE SOLICITUD DE COTIZACION") is not None else None
        codigo_sc = codigo_sc or codigo_req
        app_codigo = app_por_req_sc.get((codigo_req, codigo_sc))
        if app_codigo is None:
            raise ValueError(f"La entrega '{codigo_req} / {codigo_sc}' no tiene requerimiento correspondiente en la hoja REQUERIMIENTOS.")
        ent_por_app.setdefault(app_codigo, []).append(row)

    resultado: dict[str, bytes] = {}
    for app_codigo, req_rows in req_por_app.items():
        wb_app = openpyxl.Workbook()
        ws_r = wb_app.active
        ws_r.title = "REQUERIMIENTOS"
        ws_r.append(headers_req)
        for r in req_rows:
            ws_r.append(list(r))
        ws_e = wb_app.create_sheet("ENTREGAS")
        ws_e.append(headers_ent)
        for r in ent_por_app.get(app_codigo, []):
            ws_e.append(list(r))
        stream = BytesIO()
        wb_app.save(stream)
        resultado[app_codigo] = stream.getvalue()

    if not resultado:
        raise ValueError("No se encontraron filas válidas para importar.")
    return resultado


def _separar_formato_legacy(wb, app_por_clave: dict[str, str], hoja: str | None) -> dict[str, bytes]:
    nombre = hoja or wb.sheetnames[0]
    if nombre not in wb.sheetnames:
        raise ValueError(f"La hoja '{nombre}' no existe en el archivo")
    ws = wb[nombre]
    headers = _headers_hoja(ws)
    idx = {h: i for i, h in enumerate(headers)}
    if "Squad" not in idx:
        raise ValueError("La hoja debe incluir la columna 'Squad' para importar en modo consolidado.")
    filas_por_app: dict[str, list[tuple]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo_req = _txt(row[idx.get("COD. DEL REQ")]) if idx.get("COD. DEL REQ") is not None else None
        if not codigo_req:
            continue
        squad = _txt(row[idx["Squad"]])
        app_codigo = _resolver_app_por_squad(squad, app_por_clave)
        if app_codigo is None:
            raise ValueError(f"No se pudo resolver la aplicación del requerimiento '{codigo_req}' con Squad='{squad or ''}'.")
        filas_por_app.setdefault(app_codigo, []).append(row)

    resultado: dict[str, bytes] = {}
    for app_codigo, rows in filas_por_app.items():
        wb_app = openpyxl.Workbook()
        ws_app = wb_app.active
        ws_app.title = nombre
        ws_app.append(headers)
        for r in rows:
            ws_app.append(list(r))
        stream = BytesIO()
        wb_app.save(stream)
        resultado[app_codigo] = stream.getvalue()

    if not resultado:
        raise ValueError("No se encontraron filas válidas para importar.")
    return resultado


def _headers_hoja(ws) -> list[str]:
    fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [str(v).strip() if v is not None else "" for v in fila]


def _resolver_app_por_squad(squad: str | None, app_por_clave: dict[str, str]) -> str | None:
    if not squad:
        return None
    return app_por_clave.get(squad.strip().lower())


def _txt(valor: object) -> str | None:
    if valor is None:
        return None
    texto = str(valor).replace("\xa0", " ").strip()
    return " ".join(texto.split()) or None

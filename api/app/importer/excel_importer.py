"""Importador del Excel 'BITÁCORA GENERAL' (portado del Sistema Liquidador a Beanie).

Cada fila del Excel es un requerimiento + una entrega. Varias filas con el mismo
'COD. DEL REQ' representan distintas entregas del mismo requerimiento.
"""
from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl

from app.documents.acta_trabajo import ActaTrabajo
from app.documents.aplicativo import Aplicativo, Direccion
from app.documents.bitacora import Bitacora
from app.documents.enums import AnsResultado, EstadoEntrega, EstadoRequerimiento, TipoCosto
from app.documents.festivo import Festivo
from app.documents.orden_compra import OrdenCompra
from app.documents.persona import Persona
from app.documents.requerimiento import Entrega, Facturacion, Requerimiento, Solicitud
from app.documents.squad import Squad
from app.documents.tarifa import Tarifa
from app.services.fecha_limite import calcular_fecha_limite

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


@dataclass
class ResultadoImportacion:
    filas_procesadas: int = 0
    requerimientos_creados: int = 0
    requerimientos_actualizados: int = 0
    entregas_creadas: int = 0
    entregas_actualizadas: int = 0
    festivos_cargados: int = 0
    errores: list[str] = field(default_factory=list)


class ImportadorExcel:
    """Lee un libro de Excel y vuelca los requerimientos a MongoDB."""

    def __init__(self, aplicacion_id: str, contenido: bytes, hoja: str | None = None) -> None:
        self.aplicacion_id = aplicacion_id
        self.workbook = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        self.hoja = hoja or self._hoja_principal()
        self.anio = self._inferir_anio(self.hoja)
        self.resultado = ResultadoImportacion()
        self._cache_persona: dict = {}
        self._cache_squad: dict = {}
        self._cache_aplicativo: dict = {}
        self._cache_acta: dict = {}
        self._cache_orden: dict = {}

    # ---- ejecución ----

    async def ejecutar(self) -> ResultadoImportacion:
        await self._importar_festivos()
        for payload in self._filas_fuente():
            if not self._txt(payload.get("COD. DEL REQ")):
                continue
            try:
                await self._upsert_fila(payload)
                self.resultado.filas_procesadas += 1
            except Exception as exc:  # noqa: BLE001 - se acumula y se reporta
                self.resultado.errores.append(f"{payload.get('COD. DEL REQ')}: {exc}")
        return self.resultado

    async def previsualizar(self) -> dict:
        """Calcula el resumen de cambios sin persistir datos."""
        filas = [p for p in self._filas_fuente() if self._txt(p.get("COD. DEL REQ"))]
        codigos = sorted({self._txt(p.get("COD. DEL REQ")) for p in filas if self._txt(p.get("COD. DEL REQ"))})
        existentes = await Requerimiento.find(
            {
                "aplicacion_id": self.aplicacion_id,
                "codigo_req": {"$in": codigos},
            }
        ).to_list() if codigos else []
        por_clave_norm = {
            (self._norm_codigo(r.codigo_req), self._norm_codigo(r.solicitud.codigo_sc)): r
            for r in existentes
        }

        req_nuevos: set[str] = set()
        req_actualizados: set[str] = set()
        ent_nuevas: set[tuple] = set()
        ent_actualizadas: set[tuple] = set()
        diff_reqs: list[dict] = []
        diff_ents: list[dict] = []

        for payload in filas:
            codigo_req = self._txt(payload.get("COD. DEL REQ"))
            if not codigo_req:
                continue
            codigo_sc = self._txt(payload.get("CODIGO DE SOLICITUD DE COTIZACION")) or codigo_req
            req = por_clave_norm.get((self._norm_codigo(codigo_req), self._norm_codigo(codigo_sc)))
            if req is None:
                req_nuevos.add(f"{codigo_req}::{codigo_sc}")
                if self._fila_tiene_entrega(payload):
                    numero = self._parse_numero_entrega(payload.get("CANT. DE ENTREGAS"))
                    ent_nuevas.add((codigo_req, codigo_sc, numero))
                continue

            diffs_req = self._requerimiento_diff(req, payload)
            if diffs_req:
                clave = f"{codigo_req}::{codigo_sc}"
                req_actualizados.add(clave)
                diff_reqs.append({
                    "clave": clave,
                    "nombre": req.nombre or codigo_req,
                    "cambios": diffs_req,
                })

            if self._fila_tiene_entrega(payload):
                numero = self._parse_numero_entrega(payload.get("CANT. DE ENTREGAS"))
                existente = next((e for e in req.entregas if e.numero == numero), None)
                if existente is not None:
                    diffs_ent = self._entrega_diff(existente, payload)
                    if diffs_ent:
                        ent_actualizadas.add((codigo_req, codigo_sc, numero))
                        diff_ents.append({
                            "clave": f"{codigo_req}::SC{codigo_sc}::E{numero}",
                            "cambios": diffs_ent,
                        })
                else:
                    ent_nuevas.add((codigo_req, codigo_sc, numero))

        return {
            "filas_requerimientos": len({self._txt(f.get("COD. DEL REQ")) for f in filas if self._txt(f.get("COD. DEL REQ"))}),
            "filas_entregas": len([f for f in filas if self._fila_tiene_entrega(f)]),
            "requerimientos_nuevos": len(req_nuevos),
            "requerimientos_actualizados": len(req_actualizados),
            "entregas_nuevas": len(ent_nuevas),
            "entregas_actualizadas": len(ent_actualizadas),
            "detalle_requerimientos_nuevos": sorted(req_nuevos)[:200],
            "detalle_requerimientos_actualizados": sorted(req_actualizados)[:200],
            "detalle_entregas_nuevas": sorted([f"{r}::{sc}::E{n}" for r, sc, n in ent_nuevas])[:400],
            "detalle_entregas_actualizadas": sorted([f"{r}::{sc}::E{n}" for r, sc, n in ent_actualizadas])[:400],
            "diff_requerimientos_actualizados": diff_reqs[:200],
            "diff_entregas_actualizadas": diff_ents[:400],
        }

    async def _importar_festivos(self) -> None:
        hoja = next(
            (h for h in self.workbook.sheetnames if "festivo" in self._ident(h)), None
        )
        if hoja is None:
            return
        for fila in self.workbook[hoja].iter_rows(min_row=2, values_only=True):
            valor = fila[0] if fila else None
            fecha = self._dt(valor)
            if fecha is None:
                continue
            existe = await Festivo.find_one(
                Festivo.aplicacion_id == self.aplicacion_id, Festivo.fecha == fecha
            )
            if existe is None:
                await Festivo(aplicacion_id=self.aplicacion_id, fecha=fecha).insert()
                self.resultado.festivos_cargados += 1

    def _filas_fuente(self) -> list[dict]:
        """Obtiene filas de importación desde formato nuevo (2 hojas) o legado (1 hoja)."""
        if "REQUERIMIENTOS" in self.workbook.sheetnames and "ENTREGAS" in self.workbook.sheetnames:
            return self._filas_desde_hojas_separadas()
        if self.hoja not in self.workbook.sheetnames:
            raise ValueError(f"La hoja '{self.hoja}' no existe en el archivo")
        ws = self.workbook[self.hoja]
        headers = self._headers(ws)
        return [dict(zip(headers, fila)) for fila in ws.iter_rows(min_row=2, values_only=True)]

    def _filas_desde_hojas_separadas(self) -> list[dict]:
        ws_req = self.workbook["REQUERIMIENTOS"]
        ws_ent = self.workbook["ENTREGAS"]
        headers_req = self._headers(ws_req)
        headers_ent = self._headers(ws_ent)
        if "VALOR HORA" in headers_req:
            raise ValueError(
                "La hoja REQUERIMIENTOS no debe incluir la columna 'VALOR HORA'; ese valor se calcula automáticamente."
            )

        req_por_clave: dict[tuple[str, str], dict] = {}
        reqs_por_codigo_req: dict[str, list[dict]] = {}
        for fila in ws_req.iter_rows(min_row=2, values_only=True):
            payload_req = dict(zip(headers_req, fila))
            codigo_req = self._txt(self._valor(payload_req, "COD. DEL REQ", "COD DEL REQ", "CODIGO DEL REQ"))
            if not codigo_req:
                continue
            codigo_sc = self._txt(self._valor(payload_req, "CODIGO DE SOLICITUD DE COTIZACION", "CODIGO SOLICITUD COTIZACION")) or codigo_req
            clave_norm = (self._norm_codigo(codigo_req), self._norm_codigo(codigo_sc))
            req_por_clave[clave_norm] = payload_req
            reqs_por_codigo_req.setdefault(self._norm_codigo(codigo_req), []).append(payload_req)

        filas: list[dict] = []
        claves_con_entrega: set[tuple[str, str]] = set()
        for fila in ws_ent.iter_rows(min_row=2, values_only=True):
            payload_ent = dict(zip(headers_ent, fila))
            codigo_req = self._txt(self._valor(payload_ent, "COD. DEL REQ", "COD DEL REQ", "CODIGO DEL REQ"))
            if not codigo_req:
                if self._fila_tiene_entrega(payload_ent):
                    raise ValueError(
                        "Hay una fila en la hoja ENTREGAS sin 'COD. DEL REQ'"
                    )
                continue
            codigo_sc = self._txt(self._valor(payload_ent, "CODIGO DE SOLICITUD DE COTIZACION", "CODIGO SOLICITUD COTIZACION"))
            if not codigo_sc:
                raise ValueError(
                    f"La entrega del requerimiento '{codigo_req}' no tiene 'CODIGO DE SOLICITUD DE COTIZACION'"
                )
            clave = (self._norm_codigo(codigo_req), self._norm_codigo(codigo_sc))
            base = req_por_clave.get(clave)
            if base is None:
                candidatos = reqs_por_codigo_req.get(clave[0], [])
                if len(candidatos) == 1:
                    base = candidatos[0]
                elif len(candidatos) > 1:
                    base = next(
                        (
                            fila_req
                            for fila_req in candidatos
                            if self._norm_codigo(self._txt(self._valor(fila_req, "CODIGO DE SOLICITUD DE COTIZACION", "CODIGO SOLICITUD COTIZACION")) or self._valor(fila_req, "COD. DEL REQ", "COD DEL REQ", "CODIGO DEL REQ"))
                            == clave[1]
                        ),
                        None,
                    )
                if base is None:
                    disponibles = ", ".join(
                        filter(
                            None,
                            [
                                self._txt(self._valor(fila_req, "CODIGO DE SOLICITUD DE COTIZACION", "CODIGO SOLICITUD COTIZACION"))
                                or self._txt(self._valor(fila_req, "COD. DEL REQ", "COD DEL REQ", "CODIGO DEL REQ"))
                                for fila_req in candidatos[:5]
                            ],
                        )
                    )
                    extra = f" (disponibles: {disponibles})" if disponibles else ""
                    raise ValueError(
                        f"La entrega '{codigo_req} / {codigo_sc}' no tiene fila en la hoja REQUERIMIENTOS{extra}"
                    )
            combinado = dict(base)
            combinado.update(payload_ent)
            filas.append(combinado)
            claves_con_entrega.add(clave)

        for clave, payload_req in req_por_clave.items():
            if clave in claves_con_entrega:
                continue
            combinado = dict(payload_req)
            filas.append(combinado)
        return filas

    @staticmethod
    def _fila_tiene_entrega(payload: dict) -> bool:
        campos = (
            "CANT. DE ENTREGAS",
            "HORAS A ENTREGAR",
            "FECHA COMPROMETIDA DE ENTREGAS DLLO",
            "F REAL",
            "OBSERVACIONES",
            "MES APROBACIÓN",
            "FECHA RECEPCION DE ENTREGAS",
            "FECHA DE CARGUE ENTREGAS",
            "FECHA APROBACIÓN DE ENTREGAS",
            "FECHA EJECUCION",
            "GARANTIA",
            "ACTA DE TRABAJO ENTREGA",
            "ORDEN DE COMPRA",
            "ESTADO ENTREGAS",
            "FACTURACIÓN",
        )
        for campo in campos:
            if ImportadorExcel._txt(payload.get(campo)):
                return True
        return False

    async def _aplicar_campos_automaticos(self, req: Requerimiento) -> None:
        # Fecha límite y ANS de acta
        if req.fecha_solicitud_acta is not None:
            horas = float(req.total_horas_estimadas) if req.total_horas_estimadas else None
            req.fecha_limite = await calcular_fecha_limite(
                req.aplicacion_id, req.fecha_solicitud_acta, horas
            )
            if req.fecha_real_entrega_estimacion is not None and req.fecha_limite is not None:
                req.ans_acta = (
                    AnsResultado.CUMPLE
                    if req.fecha_real_entrega_estimacion <= req.fecha_limite
                    else AnsResultado.NO_CUMPLE
                )
            else:
                req.ans_acta = None
        else:
            req.fecha_limite = None
            req.ans_acta = None

        # ANS de entrega y porcentaje por horas
        total_estimado = req.total_horas_estimadas
        if total_estimado is not None and total_estimado > 0:
            for entrega in req.entregas:
                if entrega.horas is not None:
                    entrega.porcentaje = (
                        entrega.horas / total_estimado * 100
                    ).quantize(Decimal("0.01"))
                else:
                    entrega.porcentaje = None
        else:
            for entrega in req.entregas:
                entrega.porcentaje = None

        for entrega in req.entregas:
            if entrega.fecha_recepcion is not None and entrega.fecha_comprometida is not None:
                entrega.ans_entrega = (
                    AnsResultado.CUMPLE
                    if entrega.fecha_recepcion <= entrega.fecha_comprometida
                    else AnsResultado.NO_CUMPLE
                )
            else:
                entrega.ans_entrega = None

    def _requerimiento_diff(self, req: Requerimiento, payload: dict) -> list[dict]:
        """Retorna [{campo, antes, despues}] para campos del req que realmente cambian."""
        cambios: list[dict] = []

        def add(campo: str, antes: object, despues: object) -> None:
            cambios.append({
                "campo": campo,
                "antes": self._fmt_valor(antes),
                "despues": self._fmt_valor(despues),
            })

        def chk_text(campo: str, db_val: object, excel_val: object) -> None:
            if self._norm_text(db_val) != self._norm_text(excel_val):
                add(campo, db_val, excel_val)

        def chk_decimal(campo: str, db_val: "Decimal | None", excel_raw: object) -> None:
            excel_val = self._dec(excel_raw)
            if not self._decimal_eq(db_val, excel_val):
                add(campo,
                    str(db_val) if db_val is not None else None,
                    str(excel_val) if excel_val is not None else None)

        def chk_dt(campo: str, db_val: "datetime | None", excel_raw: object) -> None:
            excel_val = self._dt(excel_raw)
            if not self._datetime_eq(db_val, excel_val):
                add(campo,
                    db_val.strftime("%Y-%m-%d %H:%M") if db_val else None,
                    excel_val.strftime("%Y-%m-%d %H:%M") if excel_val else None)

        nombre = self._campo(payload, "NOMBRE DEL REQUERIMIENTO", "NOMBRE REQUERIMIENTO", "NOMBRE DE ACTA")
        chk_text("Nombre", req.nombre, nombre)

        estado = self._parse_estado_requerimiento(payload.get("ESTADO DE REQUERIMIENTOS"))
        chk_text("Estado", req.estado, estado)

        # SC: usar _norm_codigo para evitar falsos positivos con "123.0" vs "123"
        codigo_sc_excel = self._txt(payload.get("CODIGO DE SOLICITUD DE COTIZACION")) or req.solicitud.codigo_sc
        if self._norm_codigo(codigo_sc_excel) != self._norm_codigo(req.solicitud.codigo_sc):
            add("Código SC", req.solicitud.codigo_sc, codigo_sc_excel)

        # tipo_costo: solo comparar si Excel trae un valor; celda vacía = sin cambio
        tipo_costo_excel = self._parse_tipo_costo(payload.get("TIPO DE COSTO"))
        tipo_costo_db_norm = self._norm_text(req.solicitud.tipo_costo)
        if tipo_costo_excel is not None:
            if self._enum_value(tipo_costo_excel) != tipo_costo_db_norm:
                add("Tipo de costo", self._fmt_valor(req.solicitud.tipo_costo), self._fmt_valor(tipo_costo_excel))
        elif tipo_costo_db_norm is not None and self._txt(payload.get("TIPO DE COSTO")) is None:
            # La celda está vacía pero el DB tiene valor → el usuario la borró explícitamente
            add("Tipo de costo", self._fmt_valor(req.solicitud.tipo_costo), None)

        tecnologia = self._campo(payload, "TECNOLOGIA", "TECNOLOGÍA")
        chk_text("Tecnología", req.solicitud.tecnologia, tecnologia)

        chk_decimal("Total horas estimadas", req.total_horas_estimadas, payload.get("TOTAL HORAS ESTIMADAS"))

        chk_text("Seguimiento", req.seguimiento, payload.get("SEGUIMIENTO"))

        if "FECHA Y HORA DE SOLICITUD" in payload:
            chk_dt("Fecha y hora de solicitud", req.fecha_solicitud_acta, payload.get("FECHA Y HORA DE SOLICITUD"))

        if "ACTA DE TRABAJO" in payload:
            chk_text("Acta de trabajo", req.acta_trabajo, payload.get("ACTA DE TRABAJO"))

        return cambios

    def _requerimiento_cambia(self, req: Requerimiento, payload: dict) -> bool:
        return bool(self._requerimiento_diff(req, payload))

    def _entrega_diff(self, entrega: Entrega, payload: dict) -> list[dict]:
        """Retorna [{campo, antes, despues}] para campos de la entrega que realmente cambian."""
        cambios: list[dict] = []

        def add(campo: str, antes: object, despues: object) -> None:
            cambios.append({
                "campo": campo,
                "antes": self._fmt_valor(antes),
                "despues": self._fmt_valor(despues),
            })

        def chk_text(campo: str, db_val: object, excel_val: object) -> None:
            if self._norm_text(db_val) != self._norm_text(excel_val):
                add(campo, db_val, excel_val)

        def chk_decimal(campo: str, db_val: "Decimal | None", excel_raw: object) -> None:
            excel_val = self._dec(excel_raw)
            if not self._decimal_eq(db_val, excel_val):
                add(campo,
                    str(db_val) if db_val is not None else None,
                    str(excel_val) if excel_val is not None else None)

        def chk_dt(campo: str, db_val: "datetime | None", excel_raw: object) -> None:
            excel_val = self._dt(excel_raw)
            if not self._datetime_eq(db_val, excel_val):
                add(campo,
                    db_val.strftime("%Y-%m-%d") if db_val else None,
                    excel_val.strftime("%Y-%m-%d") if excel_val else None)

        if "HORAS A ENTREGAR" in payload:
            chk_decimal("Horas a entregar", entrega.horas, payload.get("HORAS A ENTREGAR"))

        if "FECHA COMPROMETIDA DE ENTREGAS DLLO" in payload:
            chk_dt("Fecha comprometida", entrega.fecha_comprometida, payload.get("FECHA COMPROMETIDA DE ENTREGAS DLLO"))

        if "F REAL" in payload or "FECHA RECEPCION DE ENTREGAS" in payload:
            chk_dt("F. real", entrega.fecha_recepcion,
                   self._campo(payload, "F REAL", "FECHA RECEPCION DE ENTREGAS"))

        if "GARANTIA" in payload:
            garantia_excel = self._parse_bool(payload.get("GARANTIA"))
            garantia_db = bool(entrega.garantia)
            if garantia_excel != garantia_db:
                add("Garantía", "SI" if garantia_db else "NO", "SI" if garantia_excel else "NO")

        if "ESTADO ENTREGAS" in payload:
            estado_excel = self._parse_estado_entrega(payload.get("ESTADO ENTREGAS"), entrega)
            chk_text("Estado entrega", entrega.estado, estado_excel)

        if "MES APROBACIÓN" in payload:
            chk_text("Mes aprobación", entrega.mes_aprobacion, payload.get("MES APROBACIÓN"))

        if "OBSERVACIONES" in payload:
            chk_text("Observaciones", entrega.observaciones, payload.get("OBSERVACIONES"))

        return cambios

    def _entrega_cambia(self, entrega: Entrega, payload: dict) -> bool:
        return bool(self._entrega_diff(entrega, payload))

    @staticmethod
    def _fmt_valor(valor: object) -> str | None:
        """Formatea un valor para mostrar en el diff (antes/después)."""
        if valor is None:
            return None
        if hasattr(valor, "value"):
            s = str(getattr(valor, "value")).strip()
        else:
            s = str(valor).strip()
        return s or None

    @staticmethod
    def _norm_text(valor: object) -> str | None:
        if valor is None:
            return None
        if hasattr(valor, "value"):
            valor = getattr(valor, "value")
        texto = str(valor).replace("\xa0", " ").strip()
        return " ".join(texto.split()).lower() or None

    @staticmethod
    def _enum_value(valor: object) -> str | None:
        if valor is None:
            return None
        if hasattr(valor, "value"):
            return str(getattr(valor, "value")).strip().lower()
        return str(valor).strip().lower()

    @staticmethod
    def _decimal_eq(a: Decimal | None, b: Any) -> bool:
        if a is None and b is None:
            return True
        if a is None or b is None:
            return False
        try:
            b_dec = b if isinstance(b, Decimal) else Decimal(str(b))
        except (InvalidOperation, ValueError):
            return False
        return a == b_dec

    @staticmethod
    def _datetime_eq(a: datetime | None, b: datetime | None) -> bool:
        if a is None and b is None:
            return True
        if a is None or b is None:
            return False
        a_naive = a.replace(tzinfo=None)
        b_naive = b.replace(tzinfo=None)
        return a_naive == b_naive

    def _parse_mes_facturacion(self, valor: object) -> datetime | None:
        raw = self._txt(valor)
        if not raw:
            return None
        for nombre, numero in MESES.items():
            if nombre in raw.lower():
                return datetime(self.anio, numero, 1)
        return None

    @staticmethod
    def _norm_codigo(valor: object) -> str:
        txt = ImportadorExcel._txt(valor) or ""
        if not txt:
            return ""
        # Normaliza códigos numéricos que Excel puede serializar como 123.0
        if re.fullmatch(r"\d+\.0+", txt):
            txt = txt.split(".")[0]
        return txt.strip().upper()

    # ---- fila ----

    async def _upsert_fila(self, payload: dict) -> None:
        codigo_req = self._txt(payload.get("COD. DEL REQ"))
        codigo_sc = self._txt(payload.get("CODIGO DE SOLICITUD DE COTIZACION")) or codigo_req
        req = await Requerimiento.find_one(
            Requerimiento.aplicacion_id == self.aplicacion_id,
            Requerimiento.codigo_req == codigo_req,
            Requerimiento.solicitud.codigo_sc == codigo_sc,
        )
        creado = req is None
        estado = self._parse_estado_requerimiento(payload.get("ESTADO DE REQUERIMIENTOS"))
        if req is None:
            req = Requerimiento(
                aplicacion_id=self.aplicacion_id,
                codigo_req=codigo_req,
                solicitud=Solicitud(codigo_sc=codigo_sc),
                estado=estado,
                entregas=[],
            )

        # Detectar cambios ANTES de hidratar para conteo preciso
        req_campos_cambian = creado or bool(self._requerimiento_diff(req, payload))

        await self._hidratar_solicitud(req.solicitud, payload)
        self._hidratar_requerimiento(req, payload)
        req.estado = estado

        entrega_creada = False
        entrega_actualizada = False

        if self._fila_tiene_entrega(payload):
            numero = self._parse_numero_entrega(payload.get("CANT. DE ENTREGAS"))
            entrega = next((e for e in req.entregas if e.numero == numero), None)
            if entrega is None:
                entrega = Entrega(numero=numero, estado=EstadoEntrega.PENDIENTE)
                req.entregas.append(entrega)
                entrega_creada = True
            else:
                entrega_actualizada = bool(self._entrega_diff(entrega, payload))
            await self._hidratar_entrega(entrega, payload)
            self._hidratar_facturacion(entrega, payload)
            req.entregas.sort(key=lambda e: e.numero)
            req.cantidad_entregas = len(req.entregas)

            if entrega_creada:
                self.resultado.entregas_creadas += 1
            elif entrega_actualizada:
                self.resultado.entregas_actualizadas += 1

        hay_cambios = req_campos_cambian or entrega_creada or entrega_actualizada
        if not hay_cambios:
            return  # nada que guardar

        await self._aplicar_campos_automaticos(req)
        req.marcar_actualizado()

        if creado:
            await req.insert()
            self.resultado.requerimientos_creados += 1
            await Bitacora(
                aplicacion_id=self.aplicacion_id,
                entidad_tipo="requerimiento",
                entidad_id=str(req.id),
                accion="importacion",
                descripcion=f"Requerimiento importado desde Excel (estado: {estado})",
                autor="IMPORTADOR",
            ).insert()
        else:
            await req.save()
            self.resultado.requerimientos_actualizados += 1

    async def _hidratar_solicitud(self, sol: Solicitud, payload: dict) -> None:
        sol.codigo_sc = (
            self._txt(payload.get("CODIGO DE SOLICITUD DE COTIZACION")) or sol.codigo_sc
        )
        sol.fecha_solicitud = self._dt(
            self._campo(payload, "FECHA DE SOLICITUD", "FECHA DE SOLICITUD ")
        )
        direccion = self._campo(payload, "DIRECCIÓN", "DIRECCIÓN ", "DIRECCION", "DIRECCION SOLUCION")
        sol.aplicativo_id = await self._aplicativo(
            self._campo(payload, "APLICACIÓN", "APLICACION"), direccion
        )
        sol.squad_id = await self._squad(payload.get("Squad"), payload.get("LT HITSS"))
        sol.lt_hitss_id = await self._persona(payload.get("LT HITSS"), "LT_HITSS")
        sol.lt_epm_id = await self._persona(payload.get("LT EPM"), "LT_EPM")
        sol.scrum_id = await self._persona(payload.get("SCRUM"), "SCRUM")
        sol.tipo_costo = self._parse_tipo_costo(payload.get("TIPO DE COSTO"))
        sol.estado = self._txt(payload.get("ESTADO DE REQUERIMIENTOS"))
        sol.anio_tarifa = self.anio
        sol.tecnologia = self._campo(payload, "TECNOLOGIA", "TECNOLOGÍA")
        sol.tarifa_id = await self._tarifa(payload, sol.lt_hitss_id)

    def _hidratar_requerimiento(self, req: Requerimiento, payload: dict) -> None:
        req.nombre = self._campo(
            payload,
            "NOMBRE DEL REQUERIMIENTO",
            "NOMBRE REQUERIMIENTO",
            "NOMBRE DE ACTA",
        ) or req.nombre
        req.total_horas_estimadas = self._dec(payload.get("TOTAL HORAS ESTIMADAS"))
        if "FECHA REAL ENTREGA DE ESTIMACIONES" in payload:
            req.fecha_real_entrega_estimacion = self._dt(payload.get("FECHA REAL ENTREGA DE ESTIMACIONES"))
        if "ANS ESTIMACIONES" in payload:
            req.ans_estimacion = self._parse_ans(payload.get("ANS ESTIMACIONES"))
        if "Mes" in payload:
            req.mes_objetivo = self._parse_mes(payload.get("Mes"))
        if "FECHA Y HORA DE SOLICITUD" in payload:
            req.fecha_solicitud_acta = self._dt(payload.get("FECHA Y HORA DE SOLICITUD"))
        if "ACTA DE TRABAJO" in payload:
            req.acta_trabajo = self._txt(payload.get("ACTA DE TRABAJO"))
        req.seguimiento = self._txt(payload.get("SEGUIMIENTO"))

    async def _hidratar_entrega(self, entrega: Entrega, payload: dict) -> None:
        entrega.horas = self._dec(payload.get("HORAS A ENTREGAR"))
        if "% DE ENTREGA" in payload:
            entrega.porcentaje = self._dec(payload.get("% DE ENTREGA"))
        entrega.fecha_comprometida = self._dt(payload.get("FECHA COMPROMETIDA DE ENTREGAS DLLO"))
        entrega.fecha_recepcion = self._dt(
            self._campo(payload, "F REAL", "FECHA RECEPCION DE ENTREGAS")
        )
        entrega.fecha_cargue = self._dt(payload.get("FECHA DE CARGUE ENTREGAS"))
        entrega.fecha_aprobacion = self._dt(payload.get("FECHA APROBACIÓN DE ENTREGAS"))
        entrega.fecha_ejecucion = self._dt(self._campo(payload, "FECHA EJECUCION", "FECHA EJECUCION "))
        if "ANS ENTREGAS DLLO" in payload:
            entrega.ans_entrega = self._parse_ans(payload.get("ANS ENTREGAS DLLO"))
        entrega.garantia = self._parse_bool(payload.get("GARANTIA"))
        entrega.acta_trabajo_id = await self._acta(
            self._campo(payload, "ACTA DE TRABAJO ENTREGA", "ACTA DE TRABAJO")
        )
        entrega.orden_compra_id = await self._orden(payload.get("ORDEN DE COMPRA"))
        entrega.estado = self._parse_estado_entrega(payload.get("ESTADO ENTREGAS"), entrega)
        entrega.observaciones = self._txt(payload.get("OBSERVACIONES"))
        entrega.mes_aprobacion = self._txt(payload.get("MES APROBACIÓN"))

    def _hidratar_facturacion(self, entrega: Entrega, payload: dict) -> None:
        raw = self._txt(payload.get("FACTURACIÓN"))
        if not raw:
            return
        mes = None
        for nombre, numero in MESES.items():
            if nombre in raw.lower():
                mes = datetime(self.anio, numero, 1)
                break
        from app.documents.enums import EstadoFacturacion
        entrega.facturacion = Facturacion(mes_facturacion=mes, estado=EstadoFacturacion.FACTURADA)

    # ---- get-or-create ----

    async def _persona(self, raw: object, rol: str) -> str | None:
        nombre = self._txt(raw)
        if not nombre:
            return None
        # Clave normalizada para cache (evita duplicar por diferencias de case/espacios)
        nombre_norm = " ".join(nombre.strip().lower().split())
        rol_norm = rol.strip().upper()
        clave = (nombre_norm, rol_norm)
        if clave in self._cache_persona:
            return self._cache_persona[clave]
        # Solo busca — las personas deben crearse desde la vista de Personas, no desde el importer
        # Usamos \s+ en lugar de espacios literales para tolerar dobles espacios en BD
        nombre_regex = r"\s+".join(re.escape(p) for p in nombre.split())
        rol_regex = r"\s+".join(re.escape(p) for p in rol.split())
        doc = await Persona.find_one({
            "aplicacion_id": self.aplicacion_id,
            "nombre": {"$regex": f"^{nombre_regex}$", "$options": "i"},
            "rol_operativo": {"$regex": f"^{rol_regex}$", "$options": "i"},
        })
        if doc is None:
            # Intentar sin restricción de app (la persona puede existir en otra app del mismo tenant)
            doc = await Persona.find_one({
                "nombre": {"$regex": f"^{nombre_regex}$", "$options": "i"},
                "rol_operativo": {"$regex": f"^{rol_regex}$", "$options": "i"},
            })
        if doc is None:
            self.resultado.errores.append(
                f"[AVISO] Persona '{nombre}' (rol {rol}) no encontrada — crea primero la persona en la vista de Personas."
            )
            self._cache_persona[clave] = None  # cachear para no relanzar la advertencia en filas siguientes
            return None
        self._cache_persona[clave] = str(doc.id)
        return str(doc.id)

    async def _squad(self, raw: object, raw_lt: object) -> str | None:
        nombre = self._txt(raw)
        if not nombre:
            return None
        if nombre in self._cache_squad:
            return self._cache_squad[nombre]
        lt_id = await self._persona(raw_lt, "LT_HITSS")
        doc = await Squad.find_one(
            Squad.aplicacion_id == self.aplicacion_id, Squad.nombre == nombre
        )
        if doc is None:
            doc = await Squad(
                aplicacion_id=self.aplicacion_id, nombre=nombre, lt_hitss_id=lt_id
            ).insert()
        self._cache_squad[nombre] = str(doc.id)
        return str(doc.id)

    async def _aplicativo(self, raw: object, raw_direccion: object) -> str | None:
        nombre = self._txt(raw)
        if not nombre:
            return None
        if nombre in self._cache_aplicativo:
            return self._cache_aplicativo[nombre]
        direccion = self._txt(raw_direccion)
        doc = await Aplicativo.find_one(
            Aplicativo.aplicacion_id == self.aplicacion_id, Aplicativo.nombre == nombre
        )
        if doc is None:
            doc = await Aplicativo(
                aplicacion_id=self.aplicacion_id,
                nombre=nombre,
                direccion=Direccion(nombre=direccion) if direccion else None,
            ).insert()
        elif direccion:
            actual = (
                doc.direccion.nombre.strip()
                if doc.direccion is not None and getattr(doc.direccion, "nombre", None)
                else None
            )
            if actual != direccion:
                doc.direccion = Direccion(nombre=direccion)
                doc.marcar_actualizado()
                await doc.save()
        self._cache_aplicativo[nombre] = str(doc.id)
        return str(doc.id)

    async def _acta(self, raw: object) -> str | None:
        codigo = self._txt(raw)
        if not codigo:
            return None
        if codigo in self._cache_acta:
            return self._cache_acta[codigo]
        doc = await ActaTrabajo.find_one(
            ActaTrabajo.aplicacion_id == self.aplicacion_id, ActaTrabajo.codigo == codigo
        )
        if doc is None:
            doc = await ActaTrabajo(aplicacion_id=self.aplicacion_id, codigo=codigo).insert()
        self._cache_acta[codigo] = str(doc.id)
        return str(doc.id)

    async def _orden(self, raw: object) -> str | None:
        numero = self._txt(raw)
        if not numero:
            return None
        if numero in self._cache_orden:
            return self._cache_orden[numero]
        doc = await OrdenCompra.find_one(
            OrdenCompra.aplicacion_id == self.aplicacion_id, OrdenCompra.numero == numero
        )
        if doc is None:
            doc = await OrdenCompra(aplicacion_id=self.aplicacion_id, numero=numero).insert()
        self._cache_orden[numero] = str(doc.id)
        return str(doc.id)

    async def _tarifa(self, payload: dict, lt_id: str | None) -> str | None:
        valor = self._dec(self._campo(payload, "VALOR HORA", " VALOR HORA"))
        if valor is None:
            return None
        ramificacion = self._txt(self._campo(payload, "TECNOLOGIA", "TECNOLOGÍA"))
        doc = await Tarifa.find_one(
            Tarifa.anio == self.anio,
            Tarifa.ramificacion == ramificacion,
        )
        if doc is None:
            doc = await Tarifa(
                aplicacion_id="global",
                anio=self.anio,
                valor_hora=valor,
                ramificacion=ramificacion,
            ).insert()
        else:
            doc.valor_hora = valor
            await doc.save()
        return str(doc.id)

    # ---- utilidades ----

    def _hoja_principal(self) -> str:
        for hoja in self.workbook.sheetnames:
            if "estimacion" in self._ident(hoja):
                return hoja
        return self.workbook.sheetnames[0]

    @staticmethod
    def _headers(ws) -> list[str]:
        fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        return [str(v).strip() if v is not None else "" for v in fila]

    @staticmethod
    def _campo(payload: dict, *claves: str) -> object:
        for clave in claves:
            valor = payload.get(clave)
            if valor not in (None, ""):
                return valor
        return None

    @staticmethod
    def _valor(payload: dict, *claves: str) -> object:
        """Busca una celda por nombre exacto o por nombre normalizado de columna."""
        for clave in claves:
            if clave in payload and payload.get(clave) not in (None, ""):
                return payload.get(clave)
        buscadas = {ImportadorExcel._ident(clave) for clave in claves}
        for key, value in payload.items():
            if ImportadorExcel._ident(key) in buscadas and value not in (None, ""):
                return value
        return None

    @staticmethod
    def _txt(valor: object) -> str | None:
        if valor is None:
            return None
        texto = str(valor).replace("\xa0", " ").strip()
        return " ".join(texto.split()) or None

    @staticmethod
    def _ident(valor: object) -> str:
        raw = ImportadorExcel._txt(valor) or ""
        normalizado = unicodedata.normalize("NFKD", raw)
        return "".join(c for c in normalizado.lower() if c.isalnum())

    @staticmethod
    def _dt(valor: object) -> datetime | None:
        if valor is None or valor == "":
            return None
        if isinstance(valor, datetime):
            return valor
        if isinstance(valor, date):
            return datetime(valor.year, valor.month, valor.day)
        texto = ImportadorExcel._txt(valor)
        if texto is None:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(texto, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(texto)
        except ValueError:
            return None

    @staticmethod
    def _dec(valor: object) -> Decimal | None:
        texto = ImportadorExcel._txt(valor)
        if texto is None:
            return None
        texto = texto.replace("%", "").replace(",", "")
        try:
            return Decimal(texto)
        except InvalidOperation:
            return None

    @staticmethod
    def _parse_bool(valor: object) -> bool:
        texto = ImportadorExcel._txt(valor)
        return bool(texto) and texto.lower() in {
            "si", "sí", "true", "x", "1", "garantia", "garantía"
        }

    @staticmethod
    def _parse_ans(valor: object) -> AnsResultado | None:
        texto = ImportadorExcel._txt(valor)
        if texto is None:
            return None
        try:
            return AnsResultado(texto.replace(" ", "_").upper())
        except ValueError:
            return None

    @staticmethod
    def _parse_tipo_costo(valor: object) -> TipoCosto | None:
        texto = ImportadorExcel._txt(valor)
        if texto is None:
            return None
        bajo = texto.lower()
        if bajo in {"tym", "t&m", "tym/t&m"}:
            return TipoCosto.TYM
        if bajo == "tiempo y materiales":
            return TipoCosto.TYM
        if bajo == "fijo":
            return TipoCosto.FIJO
        return None

    @staticmethod
    def _parse_estado_requerimiento(valor: object) -> str:
        texto = ImportadorExcel._txt(valor)
        if texto is None:
            return EstadoRequerimiento.ESTIMACION_EN_CURSO_POR_HITSS.value
        for candidato in EstadoRequerimiento:
            if candidato.value == texto:
                return candidato.value
        if texto == "REQUERIMIENTO SUSPENDIDO":
            return EstadoRequerimiento.REQUERIMIENTO_SUSPENDIDO_POR_EPM.value
        # Preserva estados personalizados configurados en la plataforma.
        return texto

    @staticmethod
    def _parse_estado_entrega(valor: object, entrega: Entrega) -> str:
        texto = ImportadorExcel._txt(valor)
        if texto is None:
            return EstadoEntrega.PENDIENTE.value
        if texto.startswith("APROBADA"):
            return EstadoEntrega.APROBADA.value
        for candidato in EstadoEntrega:
            if candidato.value == texto:
                return candidato.value
        if entrega.garantia:
            return EstadoEntrega.EN_GARANTIA.value
        # Preserva estados personalizados configurados en la plataforma.
        return texto

    @staticmethod
    def _parse_numero_entrega(valor: object) -> int:
        texto = ImportadorExcel._txt(valor)
        if texto is None:
            return 1
        if texto.upper() in ("UNICA", "ENTREGA UNICA", "ÚNICA"):
            return 1
        match = re.search(r"(\d+)", texto)
        return int(match.group(1)) if match else 1

    def _parse_mes(self, valor: object) -> datetime | None:
        texto = self._txt(valor)
        if texto is None:
            return None
        mes = MESES.get(texto.lower())
        return datetime(self.anio, mes, 1) if mes else None

    @staticmethod
    def _inferir_anio(hoja: str) -> int:
        match = re.search(r"(20\d{2})", hoja or "")
        return int(match.group(1)) if match else date.today().year

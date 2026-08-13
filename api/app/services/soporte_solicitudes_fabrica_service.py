"""Lógica de negocio de Soporte / Solicitudes Fábrica."""
from __future__ import annotations

import csv
import io
import logging
import re
import time
import unicodedata
from datetime import datetime, timezone
from urllib.parse import parse_qsl, quote, unquote, urlencode, urlparse, urlunparse

import httpx
import openpyxl

from app.documents.configuracion import Configuracion
from app.documents.persona import Persona
from app.documents.soporte_solicitud_fabrica import (
    ErrorValidacionSoporte,
    SoporteSolicitudFabricaSyncLog,
)
from app.documents.squad import Squad
from app.middleware.aplicacion import ContextoAplicacion
from app.repositories.soporte_solicitudes_fabrica_repository import (
    SoporteSolicitudesFabricaRepository,
)

logger = logging.getLogger(__name__)

DEFAULT_FUENTE_URL = (
    "https://globalhitss-my.sharepoint.com/:f:/r/personal/storage01_col_hitss_com/Documents/"
    "GERENCIA%20DE%20SOPORTE%20Y%20MONITOREO/12%20-%20FABRICA/DemandaSoporte/"
    "Sabana%20de%20seguimiento?csf=1&web=1&e=DB9ijh"
)
DEFAULT_ARCHIVO_NOMBRE = "Solicitudes Fabrica soporte"
DEFAULT_ARCHIVO_CANDIDATOS = (
    "Solicitudes Fabrica soporte.xlsx",
    "Solicitudes Fabrica soporte.xlsm",
)
COLUMNAS_OBJETIVO = [
    "Work Order ID",
    "Applicant",
    "Summary",
    "Detailed Description",
    "CI",
    "ASGRP",
    "Lider",
    "Squad",
    "Assigned To",
    "Status WO",
    "Priority",
    "Task ID 10",
    "Task Name 10",
    "Status Task 10",
    "Assignee Group 10",
    "Assignee 10",
    "Start Assignment Task 10",
    "End Assignment Task 10",
    "Total Hours Assigned Task 10",
    "Total Minutes Assigned Task 10",
    "Task ID 20",
    "Task Name 20",
    "Status Task 20",
    "Assignee Group 20",
    "Assignee 20",
    "Start Assignment Task 20",
    "End Assignment Task 20",
    "Total Hours Assigned Task 20",
    "Total Minutes Assigned Task 20",
    "Task ID 30",
    "Task Name 30",
    "Status Task 30",
    "Assignee Group 30",
    "Assignee 30",
    "Start Assignment Task 30",
    "End Assignment Task 30",
    "Total Hours Assigned Task 30",
    "Total Minutes Assigned Task 30",
    "Fecha_Requerida_Entrega",
    "Fecha_Requerida_Inicio",
    "Fecha_Inicio_Real",
    "Fecha_Programada_Inicio",
    "Fecha_Programada_Fin",
    "Fecha_Fin_Real",
    "Horas_Estimadas",
    "Minutos_Estimados",
    "Horas_Aprobadas",
    "Minutos_Aprobados",
    "Horas_Reales",
    "Minutos_Reales",
    "Tipo_Requerimiento_Padre",
    "Etapa de Desarrollo",
    "Fecha de Entrega a Pruebas",
    "Fecha_Asignacion_Contratista",
    "ANS_Oportunidad",
    "Estado_ANS_Oportunidad",
    "ANS_Cumplimiento",
    "Estado_ANS_Cumplimiento",
    "ANS_inicio_trabajo",
    "Estado_ANS_inicio_trabajo",
]


def _norm_txt(valor: str | None) -> str:
    raw = (valor or "").strip().lower()
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(c for c in raw if not unicodedata.combining(c))
    return " ".join(raw.split())


def _header_key(valor: str) -> str:
    t = _norm_txt(valor)
    return "".join(c for c in t if c.isalnum())

def _headers_sin_fijos(headers: list[str]) -> list[str]:
    excluidos = {"lthitss", "lider", "squad", "liderysquad", "lidersquad"}
    salida: list[str] = []
    vistos: set[str] = set()
    for h in headers:
        if not h:
            continue
        k = _header_key(h)
        if k in excluidos or h in vistos:
            continue
        vistos.add(h)
        salida.append(h)
    return salida


def _normalizar_path_sharepoint(path: str) -> str:
    """
    Convierte rutas de compartición de SharePoint (/:f:/r/..., /:x:/r/...) a ruta real (/personal/...).
    """
    p = path or ""
    m = re.match(r"^/:.\:/r(/.+)$", p, flags=re.IGNORECASE)
    if m:
        return m.group(1)
    return p


def _to_text(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat()
    return str(v).strip()


def _extraer_lider_squad(row_por_key: dict[str, str]) -> tuple[str | None, str | None]:
    # 1) columnas separadas
    lider = (row_por_key.get("lider") or row_por_key.get("lthitss") or "").strip()
    squad = (row_por_key.get("squad") or "").strip()
    if lider or squad:
        return (lider or None, squad or None)

    # 2) columna combinada "Líder y Squad"
    valor = (row_por_key.get("liderysquad") or row_por_key.get("lidersquad") or "").strip()
    if not valor:
        return (None, None)

    partes = re.split(r"\s+-\s+|\s+\|\s+|\s+/\s+|;", valor, maxsplit=1)
    if len(partes) == 2:
        return (partes[0].strip() or None, partes[1].strip() or None)
    return (None, None)


# ── Caché en memoria para listar() ──────────────────────────────────
_CACHE_TTL = 300  # 5 minutos
_cache_listar: dict[str, tuple[float, dict]] = {}

ANS_DETALLE_CAMPOS = {
    "oportunidad": {
        "levantado": "Se_levanto_ANS_Oportunidad",
        "observaciones": "Observaciones_ANS_Oportunidad",
    },
    "cumplimiento": {
        "levantado": "Se_levanto_ANS_Cumplimiento",
        "observaciones": "Observaciones_ANS_Cumplimiento",
    },
    "inicio": {
        "levantado": "Se_levanto_ANS_inicio_trabajo",
        "observaciones": "Observaciones_ANS_inicio_trabajo",
    },
}


def _cache_key(codigos: list[str]) -> str:
    return ",".join(sorted(codigos))


def _cache_invalidar() -> None:
    """Limpia toda la caché (llamar tras sincronizar)."""
    _cache_listar.clear()


class SoporteSolicitudesFabricaService:
    @staticmethod
    async def _url_fuente() -> str:
        cfg = await Configuracion.find_one(Configuracion.clave == "soporte.solicitudes_fabrica.onedrive_url")
        return (cfg.valor or "").strip() if cfg and cfg.valor else DEFAULT_FUENTE_URL

    @staticmethod
    async def _token_fuente() -> str | None:
        cfg = await Configuracion.find_one(
            Configuracion.clave == "soporte.solicitudes_fabrica.onedrive_bearer_token"
        )
        token = (cfg.valor or "").strip() if cfg and cfg.valor else ""
        return token or None

    @staticmethod
    def _build_download_candidates(url: str) -> list[str]:
        base = (url or "").strip()
        if not base:
            return []

        parsed = urlparse(base)
        if not parsed.scheme or not parsed.netloc:
            return [base]

        candidatos: list[str] = []

        def add_candidate(path: str, params: list[tuple[str, str]] | None = None) -> None:
            q = params if params is not None else list(parse_qsl(parsed.query, keep_blank_values=True))
            qs = urlencode(q, doseq=True)
            candidatos.append(
                urlunparse((parsed.scheme, parsed.netloc, path, parsed.params, qs, parsed.fragment))
            )

        original_q = list(parse_qsl(parsed.query, keep_blank_values=True))
        share_path = _normalizar_path_sharepoint(parsed.path)
        add_candidate(parsed.path, original_q)
        if share_path != parsed.path:
            add_candidate(share_path, original_q)

        q_download = [(k, v) for k, v in original_q if k.lower() not in {"web", "download", "csf"}]
        q_download.append(("download", "1"))
        add_candidate(parsed.path, q_download)
        if share_path != parsed.path:
            add_candidate(share_path, q_download)

        path_lower = share_path.lower()
        has_excel_name = any(path_lower.endswith(ext) for ext in (".xlsx", ".xlsm", ".xls"))
        if ("/:f:/" in path_lower or not has_excel_name) and not path_lower.endswith("/"):
            folder_path = share_path
        elif share_path.endswith("/"):
            folder_path = share_path[:-1]
        else:
            folder_path = share_path.rsplit("/", 1)[0]

        for nombre in DEFAULT_ARCHIVO_CANDIDATOS:
            nombre_enc = quote(nombre)
            if folder_path.endswith("/"):
                file_path = f"{folder_path}{nombre_enc}"
            else:
                file_path = f"{folder_path}/{nombre_enc}"
            add_candidate(file_path, q_download)
            source_url = f"{parsed.scheme}://{parsed.netloc}{unquote(file_path)}"
            add_candidate("/_layouts/15/download.aspx", [("SourceUrl", source_url)])

        # Elimina duplicados preservando orden
        vistos: set[str] = set()
        unicos: list[str] = []
        for c in candidatos:
            if c in vistos:
                continue
            vistos.add(c)
            unicos.append(c)
        return unicos

    @staticmethod
    async def _descargar_excel(url: str, bearer_token: str | None = None) -> bytes:
        headers = {"User-Agent": "SistemaIntegrado/soporte-sync"}
        if bearer_token:
            headers["Authorization"] = f"Bearer {bearer_token}"

        candidatos = SoporteSolicitudesFabricaService._build_download_candidates(url)
        if not candidatos:
            raise ValueError("No hay URL configurada para descargar el archivo de OneDrive.")

        ultimo_error: str | None = None
        for candidata in candidatos:
            try:
                async with httpx.AsyncClient(timeout=90.0, follow_redirects=True) as cli:
                    resp = await cli.get(candidata, headers=headers)
                    resp.raise_for_status()
                    contenido = resp.content
                    if not contenido:
                        ultimo_error = f"Archivo vacío en {candidata}"
                        continue
                    # XLSX/XLSM es ZIP (PK)
                    if contenido.startswith(b"PK"):
                        return contenido
                    # Algunas respuestas 200 retornan HTML (página de SharePoint) en lugar del archivo.
                    content_type = (resp.headers.get("content-type") or "").lower()
                    if "html" in content_type:
                        ultimo_error = f"La URL retornó HTML en vez de Excel: {candidata}"
                        continue
                    return contenido
            except httpx.HTTPStatusError as exc:
                if exc.response.status_code == 403:
                    ultimo_error = (
                        "Acceso denegado (403). La URL de OneDrive/SharePoint requiere un enlace de descarga "
                        "directa con permisos para el backend o un token Bearer válido en la configuración "
                        "'soporte.solicitudes_fabrica.onedrive_bearer_token'."
                    )
                else:
                    ultimo_error = f"HTTP {exc.response.status_code} en {candidata}"
            except httpx.HTTPError as exc:
                ultimo_error = str(exc)

        raise ValueError(f"No se pudo descargar el archivo desde OneDrive: {ultimo_error or 'sin detalle'}")

    @staticmethod
    async def _parsear_y_validar(contenido: bytes, codigos: list[str]) -> dict:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        if not wb.worksheets:
            raise ValueError("El archivo Excel no contiene hojas.")
        ws = wb.worksheets[0]

        source_headers_raw = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        source_headers = [h for h in source_headers_raw if h]
        if not source_headers:
            raise ValueError("La primera hoja no tiene encabezados.")

        idx_to_key: dict[int, str] = {}
        for i, h in enumerate(source_headers_raw):
            if h:
                idx_to_key[i] = _header_key(h)
        source_header_keys = set(idx_to_key.values())

        if not ({"lider", "squad"} <= source_header_keys) and not ({"lthitss", "squad"} <= source_header_keys) and not (
            {"liderysquad"} <= source_header_keys or {"lidersquad"} <= source_header_keys
        ):
            raise ValueError("No se encontró columna 'LT HITSS' + 'Squad' ni 'Líder y Squad' en la hoja.")

        squads = await Squad.find({"aplicacion_id": {"$in": codigos}}).to_list()

        # Los líderes técnicos pueden estar registrados en cualquier aplicación del sistema,
        # no solo en las del contexto actual. Se busca en todas las aplicaciones para evitar
        # falsos negativos cuando el mismo LT trabaja en múltiples squads/apps.
        todas_personas = await Persona.find({}).to_list()

        squad_por_nombre = {_norm_txt(s.nombre): s for s in squads if s.nombre}

        # Comparación flexible: acepta "LT_HITSS", "LT HITSS", "lt hitss", etc.
        lideres = [p for p in todas_personas if _header_key(p.rol_operativo) == "lthitss"]
        lideres_por_nombre = {_norm_txt(p.nombre): p for p in lideres if p.nombre}

        validos: list[dict] = []
        errores: list[ErrorValidacionSoporte] = []

        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            row_por_key: dict[str, str] = {}
            vacia = True
            for idx, val in enumerate(row):
                key = idx_to_key.get(idx)
                if not key:
                    continue
                txt = _to_text(val)
                row_por_key[key] = txt
                if txt:
                    vacia = False
            if vacia:
                continue

            data = {col: (row_por_key.get(_header_key(col), "") or "") for col in COLUMNAS_OBJETIVO}

            lider, squad = _extraer_lider_squad(row_por_key)
            if not lider:
                errores.append(
                    ErrorValidacionSoporte(
                        fila=excel_row,
                        lider=None,
                        squad=squad,
                        motivo="No se pudo identificar el líder en la fila.",
                    )
                )
                continue
            if not squad:
                errores.append(
                    ErrorValidacionSoporte(
                        fila=excel_row,
                        lider=lider,
                        squad=None,
                        motivo="No se pudo identificar el squad en la fila.",
                    )
                )
                continue

            lider_doc = lideres_por_nombre.get(_norm_txt(lider))
            if lider_doc is None:
                errores.append(
                    ErrorValidacionSoporte(
                        fila=excel_row,
                        lider=lider,
                        squad=squad,
                        motivo="No existe en LT_HITSS.",
                    )
                )
                continue

            squad_doc = squad_por_nombre.get(_norm_txt(squad))
            if squad_doc is None:
                errores.append(
                    ErrorValidacionSoporte(
                        fila=excel_row,
                        lider=lider,
                        squad=squad,
                        motivo="No existe en catálogo de Squad.",
                    )
                )
                continue

            squads_persona_norm = {_norm_txt(s) for s in (lider_doc.squads or [])}
            if _norm_txt(squad_doc.nombre) not in squads_persona_norm and str(squad_doc.id) not in (lider_doc.squads or []):
                errores.append(
                    ErrorValidacionSoporte(
                        fila=excel_row,
                        lider=lider,
                        squad=squad,
                        motivo="No existe la relación LT_HITSS - Squad.",
                    )
                )
                continue

            validos.append(
                {
                    "aplicacion_id": squad_doc.aplicacion_id,
                    "fila_origen": excel_row,
                    "lider": lider,
                    "squad": squad_doc.nombre,
                    "datos": data,
                }
            )

        return {
            "headers": COLUMNAS_OBJETIVO,
            "validos": validos,
            "errores": errores,
            "total_encontrados": len(validos) + len(errores),
        }

    @staticmethod
    async def listar(ctx: ContextoAplicacion) -> dict:
        key = _cache_key(ctx.codigos)
        ahora = time.monotonic()
        cached = _cache_listar.get(key)
        if cached and (ahora - cached[0]) < _CACHE_TTL:
            return cached[1]

        registros = await SoporteSolicitudesFabricaRepository.listar(ctx.codigos)
        registros_payload = []
        for r in registros:
            registros_payload.append(
                {
                    "id": str(r.id),
                    "aplicacion_id": r.aplicacion_id,
                    "fila_origen": r.fila_origen,
                    "lider": r.lider,
                    "squad": r.squad,
                    "datos": r.datos,
                    "sincronizado_en": r.sincronizado_en,
                }
            )

        log = await SoporteSolicitudFabricaSyncLog.find(
            {"$or": [{"aplicacion_id": {"$in": ctx.codigos}}, {"aplicacion_id": "__todas__"}]}
        ).sort("-iniciado_en").first_or_none()

        headers: list[str] = []
        vistos: set[str] = set()

        headers_log = _headers_sin_fijos(list(log.headers_excel or COLUMNAS_OBJETIVO)) if log else _headers_sin_fijos(COLUMNAS_OBJETIVO)
        for h in headers_log:
            if h in vistos:
                continue
            vistos.add(h)
            headers.append(h)

        for r in registros:
            for h in _headers_sin_fijos(list((r.datos or {}).keys())):
                if h in vistos:
                    continue
                vistos.add(h)
                headers.append(h)

        resultado = {
            "total": len(registros_payload),
            "ultima_actualizacion": (log.finalizado_en or log.iniciado_en) if log else None,
            "headers": headers,
            "registros": registros_payload,
        }

        _cache_listar[key] = (time.monotonic(), resultado)
        return resultado

    @staticmethod
    async def listar_paginado(
        ctx: ContextoAplicacion,
        pagina: int = 1,
        tamanio: int = 100,
        filtro_wo: str | None = None,
    ) -> dict:
        """Devuelve registros con paginación servidor para carga rápida."""
        from app.documents.soporte_solicitud_fabrica import SoporteSolicitudFabricaSyncLog

        registros, total = await SoporteSolicitudesFabricaRepository.listar_paginado(
            ctx.codigos, pagina=pagina, tamanio=tamanio, filtro_wo=filtro_wo
        )

        registros_payload = [
            {
                "id": str(r.id),
                "aplicacion_id": r.aplicacion_id,
                "fila_origen": r.fila_origen,
                "lider": r.lider,
                "squad": r.squad,
                "datos": r.datos,
                "sincronizado_en": r.sincronizado_en,
            }
            for r in registros
        ]

        log = await SoporteSolicitudFabricaSyncLog.find(
            {"$or": [{"aplicacion_id": {"$in": ctx.codigos}}, {"aplicacion_id": "__todas__"}]}
        ).sort("-iniciado_en").first_or_none()

        # Headers: usar cache si existe, sino calcular de esta página
        key = _cache_key(ctx.codigos)
        cached = _cache_listar.get(key)
        if cached:
            headers = cached[1].get("headers", [])
        else:
            headers_set: set[str] = set()
            headers_list: list[str] = []
            headers_log = _headers_sin_fijos(list(log.headers_excel or COLUMNAS_OBJETIVO)) if log else _headers_sin_fijos(COLUMNAS_OBJETIVO)
            for h in headers_log:
                if h not in headers_set:
                    headers_set.add(h)
                    headers_list.append(h)
            for r in registros:
                for h in _headers_sin_fijos(list((r.datos or {}).keys())):
                    if h not in headers_set:
                        headers_set.add(h)
                        headers_list.append(h)
            headers = headers_list

        return {
            "total": total,
            "pagina": pagina,
            "tamanio": tamanio,
            "total_paginas": (total + tamanio - 1) // tamanio,
            "ultima_actualizacion": (log.finalizado_en or log.iniciado_en) if log else None,
            "headers": headers,
            "registros": registros_payload,
        }

    @staticmethod
    async def resumen(ctx: ContextoAplicacion) -> dict:
        """Devuelve solo campos clave por registro (sin datos completos)."""
        completo = await SoporteSolicitudesFabricaService.listar(ctx)
        resumidos = []
        for r in completo.get("registros", []):
            datos = r.get("datos") or {}
            resumidos.append({
                "id": r.get("id"),
                "lider": r.get("lider"),
                "squad": r.get("squad"),
                "Work_Order_ID": datos.get("Work Order ID", ""),
                "Fecha_Fin_Real": datos.get("Fecha_Fin_Real", ""),
                "Horas_Estimadas": datos.get("Horas_Estimadas", ""),
                "Horas_Aprobadas": datos.get("Horas_Aprobadas", ""),
                "Horas_Reales": datos.get("Horas_Reales", ""),
                "Status_WO": datos.get("Status WO", ""),
                "Assigned_To": datos.get("Assigned To", ""),
                "Estado_ANS_Oportunidad": datos.get("Estado_ANS_Oportunidad", ""),
                "Estado_ANS_Cumplimiento": datos.get("Estado_ANS_Cumplimiento", ""),
                "Estado_ANS_inicio_trabajo": datos.get("Estado_ANS_inicio_trabajo", ""),
            })
        return {
            "total": completo.get("total", 0),
            "ultima_actualizacion": completo.get("ultima_actualizacion"),
            "registros": resumidos,
        }

    @staticmethod
    async def datos_ans(ctx: ContextoAplicacion) -> dict:
        """Devuelve solo campos relevantes para Detalle ANS (payload ~80% menor)."""
        _ANS_FIELDS = (
            "Work Order ID", "Fecha_Fin_Real", "Assigned To", "Status WO",
            "Estado_ANS_Oportunidad", "Estado_ANS_Cumplimiento", "Estado_ANS_inicio_trabajo",
            "Se_levanto_ANS_Oportunidad", "Observaciones_ANS_Oportunidad",
            "Se_levanto_ANS_Cumplimiento", "Observaciones_ANS_Cumplimiento",
            "Se_levanto_ANS_inicio_trabajo", "Observaciones_ANS_inicio_trabajo",
        )
        completo = await SoporteSolicitudesFabricaService.listar(ctx)
        registros_ans = []
        for r in completo.get("registros", []):
            datos = r.get("datos") or {}
            registros_ans.append({
                "id": r.get("id"),
                "lider": r.get("lider"),
                "squad": r.get("squad"),
                "datos": {k: datos.get(k, "") for k in _ANS_FIELDS},
            })
        return {"registros": registros_ans}

    @staticmethod
    async def actualizar_detalle_ans(
        ctx: ContextoAplicacion,
        registro_id: str,
        tipo: str,
        se_levanto_ans: bool | None = None,
        observaciones: str | None = None,
    ) -> dict:
        campos = ANS_DETALLE_CAMPOS.get(tipo)
        if campos is None:
            raise ValueError("Tipo de ANS no válido.")

        registro = await SoporteSolicitudesFabricaRepository.obtener(registro_id, ctx.codigos)
        if registro is None:
            raise ValueError("Registro de soporte no encontrado.")

        datos = dict(registro.datos or {})
        if se_levanto_ans is not None:
            datos[campos["levantado"]] = "SI" if se_levanto_ans else "NO"
        if observaciones is not None:
            datos[campos["observaciones"]] = observaciones.strip()

        registro.datos = datos
        await SoporteSolicitudesFabricaRepository.guardar(registro)
        _cache_invalidar()

        return {
            "id": str(registro.id),
            "aplicacion_id": registro.aplicacion_id,
            "fila_origen": registro.fila_origen,
            "lider": registro.lider,
            "squad": registro.squad,
            "datos": registro.datos,
            "sincronizado_en": registro.sincronizado_en,
        }

    @staticmethod
    async def previsualizar(
        ctx: ContextoAplicacion,
        contenido_excel: bytes,
        nombre_archivo: str | None = None,
    ) -> dict:
        if not contenido_excel:
            raise ValueError("El archivo cargado está vacío.")
        resultado = await SoporteSolicitudesFabricaService._parsear_y_validar(contenido_excel, ctx.codigos)
        errores = [
            {"fila": e.fila, "lider": e.lider, "squad": e.squad, "motivo": e.motivo}
            for e in resultado["errores"]
        ]
        return {
            "fuente_url": "archivo_local",
            "archivo": nombre_archivo or DEFAULT_ARCHIVO_NOMBRE,
            "total_encontrados": resultado["total_encontrados"],
            "registros_validos": len(resultado["validos"]),
            "registros_con_error": len(resultado["errores"]),
            "registros_que_seran_cargados": len(resultado["validos"]),
            "registros_que_no_seran_cargados": len(resultado["errores"]),
            "errores": errores,
        }

    @staticmethod
    async def sincronizar(
        ctx: ContextoAplicacion,
        contenido_excel: bytes,
        nombre_archivo: str | None = None,
    ) -> dict:
        inicio = datetime.now(timezone.utc)
        url = "archivo_local"
        if not contenido_excel:
            raise ValueError("El archivo cargado está vacío.")
        try:
            resultado = await SoporteSolicitudesFabricaService._parsear_y_validar(contenido_excel, ctx.codigos)
            filas_por_aplicacion: dict[str, list[dict]] = {codigo: [] for codigo in ctx.codigos}
            for fila in resultado["validos"]:
                app = str(fila.get("aplicacion_id") or "")
                if app in filas_por_aplicacion:
                    filas_por_aplicacion[app].append(fila)

            creados = await SoporteSolicitudesFabricaRepository.reemplazar_por_aplicacion(filas_por_aplicacion)
            _cache_invalidar()
            fin = datetime.now(timezone.utc)
            duracion_ms = int((fin - inicio).total_seconds() * 1000)
            log_app_id = "__todas__" if ctx.modo_consolidado else ctx.codigo

            log = SoporteSolicitudFabricaSyncLog(
                aplicacion_id=log_app_id,
                estado="exitoso",
                fuente_url=url,
                archivo=nombre_archivo or DEFAULT_ARCHIVO_NOMBRE,
                total_encontrados=resultado["total_encontrados"],
                validos=len(resultado["validos"]),
                con_error=len(resultado["errores"]),
                cargados=creados,
                omitidos=len(resultado["errores"]),
                duracion_ms=duracion_ms,
                headers_excel=resultado["headers"],
                iniciado_en=inicio,
                finalizado_en=fin,
                errores=resultado["errores"],
            )
            await log.insert()

            return {
                "sync_id": str(log.id),
                "total_procesados": resultado["total_encontrados"],
                "registros_creados": creados,
                "registros_omitidos": len(resultado["errores"]),
                "tiempo_ejecucion_ms": duracion_ms,
            }
        except Exception as exc:  # noqa: BLE001
            fin = datetime.now(timezone.utc)
            duracion_ms = int((fin - inicio).total_seconds() * 1000)
            logger.exception("Error sincronizando Solicitudes Fábrica")
            log_app_id = "__todas__" if ctx.modo_consolidado else ctx.codigo
            log = SoporteSolicitudFabricaSyncLog(
                aplicacion_id=log_app_id,
                estado="fallido",
                fuente_url=url,
                archivo=nombre_archivo or DEFAULT_ARCHIVO_NOMBRE,
                total_encontrados=0,
                validos=0,
                con_error=0,
                cargados=0,
                omitidos=0,
                duracion_ms=duracion_ms,
                headers_excel=[],
                iniciado_en=inicio,
                finalizado_en=fin,
                error_general=str(exc),
                errores=[],
            )
            await log.insert()
            raise

    @staticmethod
    async def descargar_errores_csv(ctx: ContextoAplicacion, sync_id: str) -> bytes:
        log = await SoporteSolicitudFabricaSyncLog.get(sync_id)
        if log is None or (log.aplicacion_id not in ctx.codigos and log.aplicacion_id != "__todas__"):
            raise ValueError("Sincronización no encontrada.")
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(["Fila", "Líder", "Squad", "Motivo"])
        for e in log.errores:
            writer.writerow([e.fila, e.lider or "", e.squad or "", e.motivo])
        return out.getvalue().encode("utf-8-sig")
